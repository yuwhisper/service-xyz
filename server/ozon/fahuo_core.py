import argparse
import base64
import io
import json
import math
import os
import random
import re
import sys
import time
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse

import pymysql
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

CROSSDOCK_DROP_OFF_WAREHOUSE_TYPES = {
    "DELIVERY_POINT",
    "SORTING_CENTER",
    "CROSS_DOCK",
    "ORDERS_RECEIVING_POINT",
    "FULL_FILLMENT",
}

# 每个供货单默认最多 30 箱；>30 时需 activate + transport/create + bind
MAX_BOXES_PER_SUPPLY_ORDER = 30
CARGOES_CREATE_BATCH_SIZE = MAX_BOXES_PER_SUPPLY_ORDER
# 合并箱唛统一输出尺寸 100mm x 100mm（等比缩放+居中，不裁切内容）
LABEL_PAGE_WIDTH_PT = 100 / 25.4 * 72
LABEL_PAGE_HEIGHT_PT = 100 / 25.4 * 72
LABEL_PAGE_SIZE_TOLERANCE_PT = 3.0
DEFAULT_SUPPLY_BOX_LIMIT = MAX_BOXES_PER_SUPPLY_ORDER
NO_PROXY = {"http": None, "https": None}

from server.ozon.config import (  # noqa: E402
    DB_CONFIG,
    DEFAULT_CROSSDOCK_DROP_OFF_NAME,
    SHIPMENT_ARCHIVE_ROOT,
    SHOP_DATA,
)

MOSCOW_CLUSTER_EXACT = "Москва, МО и Дальние регионы"
MOSCOW_CLUSTER_ALIASES = frozenset(
    {MOSCOW_CLUSTER_EXACT, "莫斯科", "莫斯科, MО и Дальние регионы"}
)


# ==================== 数据库 ====================
SHIPMENT_DETAIL_COLUMNS = """
    唯一ID, 内部订单号, 发货人, 运营发货日期, 店铺, 集群, 发货方式, 批次号, SKU, 总箱数, 单箱数量,
    单箱重量, `箱规长(cm)`, 宽, 高, 中文名称, 材料
"""

ITEM_META_KEYS = (
    "box_spec",
    "chinese_name",
    "material",
    "box_weight",
    "length",
    "width",
    "height",
)


def normalize_shipping_method(raw):
    value = (raw or "").strip()
    if value in ("直发", "DIRECT", "direct", "Direct"):
        return "直发"
    if value in ("中转", "越库", "CROSSDOCK", "crossdock", "Crossdock"):
        return "中转"
    raise ValueError(f"未知发货方式: {raw!r}，请使用「直发」或「中转」")


def _fmt_box_dim(value):
    if value is None or value == "":
        return ""
    try:
        num = float(value)
        if num == int(num):
            return str(int(num))
        return f"{num:.4f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return str(value).strip()


def format_box_spec(length, width, height):
    """箱规格式：长*宽*高（cm）。"""
    parts = [
        _fmt_box_dim(length),
        _fmt_box_dim(width),
        _fmt_box_dim(height),
    ]
    if all(parts):
        return "*".join(parts)
    return ""


def _row_detail_fields(row):
    return {
        "chinese_name": (row.get("中文名称") or "").strip(),
        "material": (row.get("材料") or "").strip(),
        "box_weight": row.get("单箱重量"),
        "length": row.get("箱规长(cm)"),
        "width": row.get("宽"),
        "height": row.get("高"),
    }


def combine_internal_order_nos(rows):
    """同一批次多个内部订单号用 + 拼接，去重保序。"""
    seen = []
    for row in rows:
        value = (row.get("内部订单号") or "").strip()
        if value and value not in seen:
            seen.append(value)
    return "+".join(seen)


def combine_shippers(rows):
    """同一批次取发货人；若多个不一致则警告并取第一个。"""
    seen = []
    for row in rows:
        value = (row.get("发货人") or "").strip()
        if value and value not in seen:
            seen.append(value)
    if len(seen) > 1:
        print(f"⚠️ 批次存在多个发货人 {seen}，文件夹命名使用: {seen[0]!r}")
    return seen[0] if seen else "未知发货人"


def sanitize_folder_name(name):
    text = (name or "").strip()
    for ch in '\\/:*?"<>|':
        text = text.replace(ch, "_")
    return text or "未知"


def cluster_folder_label(clusters):
    """单集群用集群名；多集群合并单用「多集群」。"""
    unique = [c for c in clusters if c]
    if len(unique) <= 1:
        return unique[0] if unique else "未知集群"
    return "多集群"


def is_moscow_cluster(cluster_name):
    """直发模式仅发往莫斯科集群。"""
    name = (cluster_name or "").strip()
    if not name:
        return False
    if name in MOSCOW_CLUSTER_ALIASES:
        return True
    return name.startswith("Москва") or name.startswith("莫斯科")


def build_order_output_dir(shipper, shop, batch_or_order, archive_date=None):
    """
    输出目录：{存档根}/{YYYY-MM-DD}+{发货人}+{店铺}+{批次号或订单号}
    同一单号下直发/中转合并输出总箱唛与 Excel。
    """
    if archive_date is None:
        archive_date = datetime.now().date()
    elif isinstance(archive_date, datetime):
        archive_date = archive_date.date()
    date_str = archive_date.strftime("%Y-%m-%d")
    shipper_text = sanitize_folder_name(shipper or "未知发货人")
    shop_text = sanitize_folder_name(shop or "未知店铺")
    tag_text = sanitize_folder_name(batch_or_order or "未知单号")
    folder_name = f"{date_str}+{shipper_text}+{shop_text}+{tag_text}"
    return os.path.join(SHIPMENT_ARCHIVE_ROOT, folder_name)


def build_excel_report_filenames(batch_or_order, total_boxes):
    """询价表/顺序表文件名：单号与文件夹一致，询价表含慢线与总箱数。"""
    tag = sanitize_folder_name(batch_or_order or "未知单号")
    inquiry = f"{tag}慢线{int(total_boxes)}.xlsx"
    order_sheet = f"{tag}箱唛顺序表.xlsx"
    return inquiry, order_sheet


def item_meta_kwargs(item):
    meta = {key: item.get(key) for key in ITEM_META_KEYS}
    if item.get("row_id") is not None:
        meta["row_id"] = item["row_id"]
    return meta


def sort_shipment_items_by_row_id(shipment_items):
    """按登记表唯一ID升序排列（多集群全局顺序基准）。"""
    return sorted(
        shipment_items,
        key=lambda row: (
            row.get("row_id") if row.get("row_id") is not None else 0,
            row.get("offer_id") or "",
        ),
    )


def _row_unique_id(row):
    value = row.get("唯一ID")
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def build_items_from_rows(rows):
    """
    按登记表唯一ID顺序逐行生成 ITEMS（不跨行聚合）。
    相同 SKU 多行时，箱唛/顺序表/询价表顺序与唯一ID升序一致。
    """
    items = []
    for row in sorted(rows, key=_row_unique_id):
        sku = (row.get("SKU") or "").strip()
        box_count = int(row.get("总箱数") or 0)
        per_box = int(row.get("单箱数量") or 0)
        if not sku:
            raise ValueError(f"存在空 SKU 行: {row}")
        if box_count <= 0 or per_box <= 0:
            raise ValueError(
                f"SKU {sku} 总箱数/单箱数量无效: 总箱数={box_count}, 单箱数量={per_box}"
            )
        spec = format_box_spec(
            row.get("箱规长(cm)"),
            row.get("宽"),
            row.get("高"),
        )
        items.append(
            {
                "sku": sku,
                "quantity": box_count * per_box,
                "number": per_box,
                "box_spec": spec,
                "row_id": _row_unique_id(row),
                "cluster": (row.get("集群") or "").strip(),
                **_row_detail_fields(row),
            }
        )
    return items


def build_cluster_items_map(items):
    """按集群名分组 ITEMS，保持组内唯一ID顺序。"""
    by_cluster = defaultdict(list)
    for item in items:
        cluster = (item.get("cluster") or "").strip() or "未知集群"
        by_cluster[cluster].append(item)
    return dict(by_cluster)


def combine_batch_nos(rows):
    seen = []
    for row in rows:
        value = str(row.get("批次号") or "").strip()
        if value and value not in seen:
            seen.append(value)
    return "+".join(seen) if seen else ""


def resolve_batch_or_order_key(row):
    """分组/目录用：有批次号取批次号，否则取内部订单号（订单号必有）。"""
    batch = str(row.get("批次号") or "").strip()
    if batch:
        return batch
    order_no = (row.get("内部订单号") or "").strip()
    if order_no:
        return order_no
    raise ValueError(
        f"唯一ID={row.get('唯一ID')!r} 缺少批次号与内部订单号，无法分组"
    )


def resolve_batch_or_order_from_rows(rows):
    """从多行解析文件夹/分包用的批次号或订单号。"""
    keys = []
    for row in rows or []:
        keys.append(resolve_batch_or_order_key(row))
    unique = list(dict.fromkeys(keys))
    if len(unique) > 1:
        print(
            f"⚠️ 多行批次/订单号不一致 {unique}，文件夹命名使用: {unique[0]!r}"
        )
    return unique[0] if unique else "未知单号"


def count_boxes_in_items(items):
    total = 0
    for item in items:
        per_box = int(item["number"])
        qty = int(item["quantity"])
        if per_box <= 0 or qty % per_box != 0:
            raise ValueError(
                f"SKU {item['sku']} 数量 {qty} 无法按单箱 {per_box} 整除"
            )
        total += qty // per_box
    return total


def split_items_by_box_limit(items, max_boxes=MAX_BOXES_PER_SUPPLY_ORDER):
    """
    按箱数拆分为多个 ITEMS 子集，每子集不超过 max_boxes 箱。
    同一 SKU 可跨子单拆分。
    """
    box_units = []
    for item in items:
        sku = item["sku"]
        per_box = int(item["number"])
        qty = int(item["quantity"])
        if per_box <= 0 or qty % per_box != 0:
            raise ValueError(
                f"SKU {sku} 数量 {qty} 无法按单箱 {per_box} 整除"
            )
        for _ in range(qty // per_box):
            box_units.append((sku, per_box))

    if not box_units:
        raise ValueError("ITEMS 为空，无法拆分")

    chunks = []
    for start in range(0, len(box_units), max_boxes):
        part = box_units[start : start + max_boxes]
        aggregated = defaultdict(int)
        for sku, per_box in part:
            aggregated[(sku, per_box)] += per_box
        chunks.append(
            [
                {"sku": sku, "quantity": qty, "number": per_box}
                for (sku, per_box), qty in aggregated.items()
            ]
        )
    return chunks


def fetch_pending_shipment_groups(db_config=None, ship_date=None):
    """
    读取今日待发货登记，按 日期+店铺+发货人+发货方式+批次号/订单号 分组。
    有批次号用批次号，否则用内部订单号；直发与中转分别创单；直发组内仅保留莫斯科集群行。
    """
    cfg = db_config or DB_CONFIG
    if ship_date is None:
        date_clause = "DATE(运营发货日期) = CURRENT_DATE"
        params = ()
    else:
        date_clause = "DATE(运营发货日期) = %s"
        params = (ship_date,)
    pending_sql = f"""
        SELECT {SHIPMENT_DETAIL_COLUMNS}
        FROM ods_ozon_装箱发货登记表
        WHERE {date_clause}
          AND (发货状态 IS NULL OR 发货状态 = '')
        ORDER BY 店铺, 发货人, 唯一ID
    """
    conn = pymysql.connect(**cfg)
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(pending_sql, params)
            raw_rows = cursor.fetchall()
    finally:
        conn.close()

    groups = {}
    for row in raw_rows:
        ship_date = row.get("运营发货日期")
        if isinstance(ship_date, datetime):
            ship_date = ship_date.date()
        shop = (row.get("店铺") or "").strip()
        shipper = (row.get("发货人") or "").strip() or "未知发货人"
        try:
            method = normalize_shipping_method(row.get("发货方式"))
            batch_or_order = resolve_batch_or_order_key(row)
        except ValueError as e:
            print(f"⚠️ 跳过无效行: {e}")
            continue
        key = (ship_date, shop, shipper, method, batch_or_order)
        if key not in groups:
            groups[key] = {
                "ship_date": ship_date,
                "shop": shop,
                "shipper": shipper,
                "shipping_method": method,
                "batch_or_order": batch_or_order,
                "bundle_key": (ship_date, shop, shipper, batch_or_order),
                "rows": [],
            }
        groups[key]["rows"].append(row)

    result = []
    for meta in groups.values():
        rows = meta["rows"]
        if meta["shipping_method"] == "直发":
            moscow_rows = [
                row for row in rows if is_moscow_cluster(row.get("集群"))
            ]
            skipped = len(rows) - len(moscow_rows)
            if skipped:
                print(
                    f"ℹ️ [{meta['shop']}/{meta['shipper']}] 直发单跳过 "
                    f"{skipped} 行非莫斯科集群"
                )
            if not moscow_rows:
                print(
                    f"⚠️ [{meta['shop']}/{meta['shipper']}] 直发组无莫斯科行，跳过"
                )
                continue
            rows = moscow_rows
            meta["rows"] = rows

        meta["items"] = build_items_from_rows(rows)
        meta["cluster_items"] = build_cluster_items_map(meta["items"])
        meta["clusters"] = sorted(meta["cluster_items"].keys())
        meta["cluster"] = meta["clusters"][0] if len(meta["clusters"]) == 1 else ""
        meta["batch_no"] = combine_batch_nos(rows)
        meta["internal_order_no"] = combine_internal_order_nos(rows)
        result.append(meta)
    result.sort(
        key=lambda g: (g["shop"], g["shipper"], 0 if g["shipping_method"] == "直发" else 1)
    )
    return result


def fetch_batch_rows(shop, batch_no, ship_date=None, db_config=None):
    """按店铺+批次号读取登记明细行（断点续传用，不限制发货状态）。"""
    cfg = db_config or DB_CONFIG
    sql = f"""
        SELECT {SHIPMENT_DETAIL_COLUMNS}
        FROM ods_ozon_装箱发货登记表
        WHERE 店铺 = %s AND 批次号 = %s
    """
    params = [shop, batch_no]
    if ship_date:
        sql += " AND DATE(运营发货日期) = DATE(%s)"
        params.append(ship_date)
    sql += " ORDER BY 唯一ID"
    conn = pymysql.connect(**cfg)
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()
    finally:
        conn.close()
    if not rows:
        raise ValueError(f"未找到店铺={shop!r} 批次={batch_no!r} 的登记记录")
    return rows


def fetch_batch_items(shop, batch_no, ship_date=None, db_config=None):
    return build_items_from_rows(
        fetch_batch_rows(shop, batch_no, ship_date=ship_date, db_config=db_config)
    )


def fetch_batch_data(shop, batch_no, ship_date=None, db_config=None):
    rows = fetch_batch_rows(
        shop, batch_no, ship_date=ship_date, db_config=db_config
    )
    return build_items_from_rows(rows), combine_internal_order_nos(rows), combine_shippers(rows)


def mark_rows_shipment_applied(rows, db_config=None):
    """成功后按登记表唯一ID回写发货状态。"""
    row_ids = []
    for row in rows or []:
        row_id = row.get("唯一ID") if isinstance(row, dict) else row
        if row_id is None or row_id == "":
            continue
        row_ids.append(int(row_id))
    row_ids = list(dict.fromkeys(row_ids))
    if not row_ids:
        return 0

    cfg = db_config or DB_CONFIG
    status_text = "已申请"
    placeholders = ", ".join(["%s"] * len(row_ids))
    conn = pymysql.connect(**cfg)
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                UPDATE ods_ozon_装箱发货登记表
                SET 发货状态 = %s
                WHERE 唯一ID IN ({placeholders})
                  AND (发货状态 IS NULL OR 发货状态 = '')
                """,
                [status_text, *row_ids],
            )
            affected = cursor.rowcount
        conn.commit()
        return affected
    finally:
        conn.close()


def mark_batch_applied(ship_date, shop, cluster, batch_no, order_id, db_config=None):
    """兼容旧调用：按批次条件回写发货状态。"""
    del order_id
    cfg = db_config or DB_CONFIG
    status_text = "已申请"
    conn = pymysql.connect(**cfg)
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                UPDATE ods_ozon_装箱发货登记表
                SET 发货状态 = %s
                WHERE DATE(运营发货日期) = DATE(%s)
                  AND 店铺 = %s
                  AND 集群 = %s
                  AND 批次号 = %s
                  AND (发货状态 IS NULL OR 发货状态 = '')
                """,
                (status_text, ship_date, shop, cluster, batch_no),
            )
            affected = cursor.rowcount
        conn.commit()
        return affected
    finally:
        conn.close()


# ==================== 通用抗频限请求核心 ====================
def ozon_post(
    url, headers, payload, proxies=None, max_retries=5, initial_delay=2
):
    """带自动重试机制的请求器，专门应对 Ozon 的 429 频限拦截"""
    retry_delay = initial_delay
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.post(
                url, headers=headers, json=payload, proxies=proxies
            )
            if response.status_code == 429:
                print(
                    f"⚠️ 触发 Ozon 频率限制 (429)。正在等待 {retry_delay} 秒后进行第 {attempt}/{max_retries} 次重试..."
                )
                time.sleep(retry_delay)
                retry_delay *= 2
                continue
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            print(
                f"❌ Ozon API 返回错误码 [{response.status_code}]: {response.text}"
            )
            raise e
        except Exception as e:
            print(f"❌ 网络或系统异常: {e}")
            raise e
    raise Exception(
        f"❌ 在重试 {max_retries} 次后，依然由于 429 频限无法完成请求。"
    )


def ozon_download_binary(
    url,
    headers,
    proxies=None,
    max_retries=5,
    initial_delay=2,
    method="GET",
    json_body=None,
):
    """下载 Ozon 返回的二进制文件（箱唛 zip/pdf 等）。"""
    retry_delay = initial_delay
    download_headers = {
        k: v for k, v in headers.items() if k.lower() != "content-type"
    }
    download_headers.setdefault(
        "Accept",
        "application/pdf, application/zip, application/octet-stream, */*",
    )
    for attempt in range(1, max_retries + 1):
        response = None
        try:
            if method.upper() == "POST":
                response = requests.post(
                    url,
                    headers=download_headers,
                    json=json_body,
                    proxies=proxies,
                )
            else:
                response = requests.get(
                    url, headers=download_headers, proxies=proxies
                )
            if response.status_code == 429:
                print(
                    f"⚠️ 下载触发频率限制 (429)。等待 {retry_delay} 秒后重试 ({attempt}/{max_retries})..."
                )
                time.sleep(retry_delay)
                retry_delay *= 2
                continue
            if response.status_code == 404:
                raise requests.exceptions.HTTPError(
                    f"404 for {method} {url}", response=response
                )
            response.raise_for_status()
            content_type = response.headers.get("Content-Type", "")
            if "application/json" in content_type.lower():
                data = response.json()
                file_url = (
                    data.get("file_url")
                    or data.get("url")
                    or (data.get("result") or {}).get("file_url")
                    or (data.get("result") or {}).get("url")
                )
                if file_url:
                    return ozon_download_binary(
                        file_url, headers, proxies, max_retries, initial_delay
                    )
                raise ValueError(f"下载接口返回 JSON 但无文件地址: {data}")
            filename = None
            content_disposition = response.headers.get("Content-Disposition", "")
            if "filename=" in content_disposition:
                filename = (
                    content_disposition.split("filename=", 1)[-1]
                    .strip('"\' ')
                )
            return response.content, content_type, filename
        except requests.exceptions.HTTPError as e:
            status = response.status_code if response is not None else "?"
            body = response.text[:500] if response is not None else str(e)
            print(f"❌ 下载失败 [{status}] {method} {url}: {body}")
            raise e
        except Exception as e:
            print(f"❌ 下载网络异常: {e}")
            raise e
    raise Exception(f"❌ 下载在重试 {max_retries} 次后仍失败。")


# ==================== 基础数据解析辅助 ====================
def build_cargoes_get_payload(supply_id):
    return {"supply_ids": [int(supply_id)]}


def parse_cargoes_get_response(get_res):
    supplies = (
        get_res.get("supply") or get_res.get("result", {}).get("supply") or []
    )
    all_cargoes = []
    for supply in supplies:
        all_cargoes.extend(supply.get("cargoes") or [])
    if not all_cargoes:
        all_cargoes = (
            get_res.get("cargoes")
            or get_res.get("result", {}).get("cargoes")
            or []
        )
    return all_cargoes


def fetch_cargoes_list(supply_id, headers, proxies):
    get_res = ozon_post(
        "https://api-seller.ozon.ru/v1/cargoes/get",
        headers,
        build_cargoes_get_payload(supply_id),
        proxies,
    )
    return parse_cargoes_get_response(get_res)


def extract_cargo_ids_from_get(get_res):
    if get_res.get("supply") or get_res.get("result", {}).get("supply"):
        cargoes = parse_cargoes_get_response(get_res)
    else:
        cargoes = (
            get_res.get("cargoes")
            or get_res.get("result", {}).get("cargoes")
            or []
        )
    cargo_ids = []
    for cargo in cargoes:
        cargo_id = cargo.get("cargo_id") or cargo.get("id")
        if cargo_id:
            cargo_ids.append(int(cargo_id))
    return cargo_ids


def extract_ordered_cargo_ids_from_create_info(info_res):
    result = info_res.get("result") or {}
    cargoes = result.get("cargoes") or info_res.get("cargoes") or []
    pairs = []
    for cargo in cargoes:
        if not isinstance(cargo, dict):
            continue
        key = cargo.get("key") or ""
        value = cargo.get("value")
        if isinstance(value, dict):
            cargo_id = value.get("cargo_id")
        else:
            cargo_id = cargo.get("cargo_id")
        if cargo_id:
            pairs.append((key, int(cargo_id)))
    pairs.sort(key=lambda item: item[0])
    return [cargo_id for _, cargo_id in pairs]


def fetch_new_box_cargo_ids(supply_id, exclude_ids, limit, headers, proxies):
    """从 supply 中取尚未使用的 BOX cargo_id（按 id 排序）。"""
    exclude = set(exclude_ids or [])
    box_ids = []
    for cargo in fetch_cargoes_list(supply_id, headers, proxies):
        if (cargo.get("type") or "").upper() != "BOX":
            continue
        cargo_id = cargo.get("cargo_id") or cargo.get("id")
        if cargo_id:
            cid = int(cargo_id)
            if cid not in exclude:
                box_ids.append(cid)
    box_ids.sort()
    return box_ids[:limit]


def extract_label_file_info(label_get_res):
    result = label_get_res.get("result") or {}
    file_url = result.get("file_url") or label_get_res.get("file_url")
    file_guid = result.get("file_guid") or label_get_res.get("file_guid")
    file_content_b64 = (
        result.get("file_content")
        or result.get("content")
        or label_get_res.get("file_content")
    )
    return file_url, file_guid, file_content_b64


def decode_label_file_content(file_content_b64):
    if not file_content_b64:
        return None
    if isinstance(file_content_b64, str):
        return base64.b64decode(file_content_b64)
    return file_content_b64


def download_label_from_url(file_url, headers, proxies):
    print(f"📥 从 file_url 下载箱唛: {file_url}")
    try:
        return ozon_download_binary(file_url, headers, proxies)
    except requests.exceptions.HTTPError:
        print("⚠️ 带 Seller API 头下载失败，尝试直连 CDN...")
        plain_headers = {
            "Accept": "application/pdf, application/zip, application/octet-stream, */*"
        }
        return ozon_download_binary(file_url, plain_headers, proxies)


def download_cargoes_label_file(file_guid, headers, proxies):
    base = "https://api-seller.ozon.ru/v1/cargoes-label/file"
    strategies = [
        ("GET", f"{base}/{file_guid}", None),
        ("POST", base, {"file_guid": file_guid}),
        ("GET", f"{base}?file_guid={file_guid}", None),
    ]
    last_error = None
    for wait_sec in (0, 2, 4):
        if wait_sec:
            print(f"⏳ 等待 {wait_sec} 秒后重试下载箱唛...")
            time.sleep(wait_sec)
        for method, url, body in strategies:
            try:
                print(f"📥 尝试 {method} {url}")
                return ozon_download_binary(
                    url, headers, proxies, method=method, json_body=body
                )
            except requests.exceptions.HTTPError as e:
                last_error = e
                if e.response is not None and e.response.status_code == 404:
                    continue
                raise
    if last_error:
        raise last_error
    raise Exception("❌ 无法下载箱唛文件")


def download_box_label_content(
    file_url, file_guid, file_content_b64, headers, proxies
):
    embedded = decode_label_file_content(file_content_b64)
    if embedded:
        print("📥 使用 get 响应中的 file_content（Base64）")
        return embedded, "application/pdf", None

    if file_url:
        return download_label_from_url(file_url, headers, proxies)

    if file_guid:
        print("⚠️ 无 file_url，回退已废弃的 file_guid 下载接口...")
        return download_cargoes_label_file(file_guid, headers, proxies)

    raise ValueError("箱唛 get 响应中无 file_url / file_content / file_guid")


def _normalize_label_text(text):
    return re.sub(r"\s+", "", text or "")


def _page_cargo_id(page, expected_set):
    compact = _normalize_label_text(page.get_text())
    matched = [cid for cid in expected_set if str(cid) in compact]
    if not matched:
        return None
    if len(matched) == 1:
        return matched[0]
    return max(matched, key=lambda cid: len(str(cid)))


def _is_standard_label_page(rect):
    return (
        abs(rect.width - LABEL_PAGE_WIDTH_PT) <= LABEL_PAGE_SIZE_TOLERANCE_PT
        and abs(rect.height - LABEL_PAGE_HEIGHT_PT) <= LABEL_PAGE_SIZE_TOLERANCE_PT
    )


def _append_normalized_label_page(dst_doc, src_doc, src_page_idx):
    """
    将源 PDF 的一页写入目标文档。
    非标准尺寸（如直发 A4、中转 120×75）等比缩放至 100mm×100mm，居中放置，不拉伸变形。
    """
    import fitz

    src_page = src_doc[src_page_idx]
    src_rect = src_page.rect
    if _is_standard_label_page(src_rect):
        dst_doc.insert_pdf(src_doc, from_page=src_page_idx, to_page=src_page_idx)
        return

    target_w = LABEL_PAGE_WIDTH_PT
    target_h = LABEL_PAGE_HEIGHT_PT
    scale = min(target_w / src_rect.width, target_h / src_rect.height)
    scaled_w = src_rect.width * scale
    scaled_h = src_rect.height * scale
    x0 = (target_w - scaled_w) / 2
    y0 = (target_h - scaled_h) / 2
    dest_rect = fitz.Rect(x0, y0, x0 + scaled_w, y0 + scaled_h)
    new_page = dst_doc.new_page(width=target_w, height=target_h)
    new_page.show_pdf_page(dest_rect, src_doc, src_page_idx)


def reorder_cargoes_label_pdf(file_bytes, ordered_cargo_ids):
    try:
        import fitz
    except ImportError as e:
        raise ImportError(
            "需要安装 PyMuPDF (pip install pymupdf) 才能重排箱唛页序"
        ) from e

    src = fitz.open(stream=file_bytes, filetype="pdf")
    page_by_cargo = {}
    expected_set = set(ordered_cargo_ids)

    for page_idx in range(src.page_count):
        cid = _page_cargo_id(src[page_idx], expected_set)
        if cid is not None and cid not in page_by_cargo:
            page_by_cargo[cid] = page_idx

    missing = [cid for cid in ordered_cargo_ids if cid not in page_by_cargo]
    if missing:
        src.close()
        raise ValueError(
            f"PDF 中未找到 {len(missing)} 个货位的页面，无法重排: {missing[:3]}..."
        )

    dst = fitz.open()
    for cid in ordered_cargo_ids:
        page_idx = page_by_cargo[cid]
        _append_normalized_label_page(dst, src, page_idx)

    ordered_bytes = dst.tobytes()
    dst.close()
    src.close()
    return ordered_bytes


def build_box_meta_from_items(shipment_items):
    """按 ITEMS 拆箱顺序生成每箱明细，与 cargoes_list 一一对应。"""
    meta = []
    for row in shipment_items:
        offer_id = row["offer_id"]
        total_qty = row["quantity"]
        per_box = row["number"]
        box_count = total_qty // per_box
        for _ in range(box_count):
            meta.append(
                {
                    "sku": offer_id,
                    "cluster": row.get("cluster") or "",
                    "row_id": row.get("row_id"),
                    "box_spec": row.get("box_spec") or "",
                    "chinese_name": row.get("chinese_name") or "",
                    "material": row.get("material") or "",
                    "box_weight": row.get("box_weight"),
                    "length": row.get("length"),
                    "width": row.get("width"),
                    "height": row.get("height"),
                    "per_box_qty": per_box,
                }
            )
    return meta


def format_cargo_id(value):
    """长 cargo_id 必须以字符串保存，避免 Excel 浮点精度丢失。"""
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        text = value.strip()
        if text.endswith(".0"):
            text = text[:-2]
        return text
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return f"{value:.0f}"
    return str(value)


def _set_text_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value="" if value is None else value)
    cell.number_format = "@"
    return cell


def _apply_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def enrich_box_meta_with_cargo_ids(box_meta, cargo_ids):
    enriched = []
    for idx, row in enumerate(box_meta):
        item = dict(row)
        if idx < len(cargo_ids):
            item["cargo_id"] = format_cargo_id(cargo_ids[idx])
        enriched.append(item)
    return enriched


def fetch_product_details(skus, db_config=None):
    """从聚水潭 API 批量查询图片 URL（pic）与运营成本价（other_price_5）。"""
    del db_config  # 保留参数兼容旧调用，已不再读 MySQL 商品资料表
    from server.jushuitan.client import query_skus

    unique_skus = list(dict.fromkeys(s for s in skus if s))
    if not unique_skus:
        return {}
    return query_skus(unique_skus)


def _excel_num(value):
    if value is None or value == "":
        return None
    try:
        num = float(value)
        if num == int(num):
            return int(num)
        return num
    except (TypeError, ValueError):
        return value


def download_product_image(url, cache_dir, sku):
    if not url:
        return None
    os.makedirs(cache_dir, exist_ok=True)
    path_part = urlparse(url).path
    ext = os.path.splitext(path_part)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"):
        ext = ".jpg"
    safe_sku = re.sub(r"[^\w\-]", "_", sku)
    path = os.path.join(cache_dir, f"{safe_sku}{ext}")
    if os.path.isfile(path) and os.path.getsize(path) > 0:
        return path
    try:
        resp = requests.get(
            url,
            timeout=30,
            proxies=NO_PROXY,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        resp.raise_for_status()
        with open(path, "wb") as f:
            f.write(resp.content)
        return path
    except requests.RequestException as exc:
        print(f"⚠️ SKU {sku} 图片下载失败: {exc}")
        return None


def prepare_product_image_paths(product_details, cache_dir):
    """按 SKU 预下载商品图片，避免每箱重复请求。"""
    image_paths = {}
    for sku, info in product_details.items():
        url = (info or {}).get("image_url")
        if not url:
            continue
        path = download_product_image(url, cache_dir, sku)
        if path:
            image_paths[sku] = path
    return image_paths


def export_order_sheet_xlsx(output_path, box_meta, internal_order_no):
    wb = Workbook()
    ws = wb.active
    ws.title = "顺序表"
    ws.append(
        ["聚水潭单号", "货代贴标顺序", "sku", "单箱数量", "规格", "", "", "俄罗斯箱唛"]
    )
    ws.merge_cells("E1:G1")
    ws["E1"].alignment = Alignment(horizontal="center", vertical="center")

    data_start = 2
    for idx, row in enumerate(box_meta, start=1):
        excel_row = data_start + idx - 1
        ws.cell(row=excel_row, column=1, value=internal_order_no if idx == 1 else "")
        ws.cell(row=excel_row, column=2, value=f"第{idx}张")
        ws.cell(row=excel_row, column=3, value=row.get("sku") or "")
        ws.cell(row=excel_row, column=4, value=_excel_num(row.get("per_box_qty")))
        ws.cell(row=excel_row, column=5, value=_excel_num(row.get("length")))
        ws.cell(row=excel_row, column=6, value=_excel_num(row.get("width")))
        ws.cell(row=excel_row, column=7, value=_excel_num(row.get("height")))
        _set_text_cell(ws, excel_row, 8, format_cargo_id(row.get("cargo_id")))

    if box_meta:
        last_row = data_start + len(box_meta) - 1
        if last_row > data_start:
            ws.merge_cells(f"A{data_start}:A{last_row}")
        ws[f"A{data_start}"].alignment = Alignment(
            vertical="top", wrap_text=True
        )

    _apply_col_widths(
        ws,
        {
            "A": 14.0,
            "B": 18.0,
            "C": 18.0,
            "D": 10.0,
            "E": 8.0,
            "F": 8.0,
            "G": 8.0,
            "H": 22.0,
        },
    )
    wb.save(output_path)


def _setup_inquiry_template(ws):
    """按 159416 询价表模板搭建表头、合并单元格与列宽。"""
    title = "义乌市供应链管理有限公司入库/结算单（运费总计：$）"
    ws.append([title] + [None] * 20)
    ws.merge_cells("A1:U1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 53

    row2 = [None] * 21
    row2[0] = "票号："
    row2[3] = "运费合计（$）："
    row2[7] = "发货人："
    row2[11] = "收货人、地址及电话："
    row2[16] = "是否核销:"
    row2[19] = "发货日期："
    ws.append(row2)
    ws.merge_cells("A2:C2")
    ws.merge_cells("D2:G2")
    ws.merge_cells("H2:K2")
    ws.merge_cells("L2:P2")
    ws.merge_cells("Q2:S2")
    ws.merge_cells("T2:U2")
    ws.row_dimensions[2].height = 36

    ws.append(
        [
            "SKU",
            "品 名",
            "照片",
            "配比",
            "材 质",
            "件数",
            "数 量",
            "重 量",
            "箱规",
            None,
            None,
            "运价($)",
            "总数量",
            "总重量（KG)",
            "立 方",
            "单价（￥）",
            "货值（￥)",
            "保值($)",
            "保险费（$）",
            "运费（$)",
            "聚水潭单号",
        ]
    )
    ws.merge_cells("I3:K3")
    ws["I3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 27

    _apply_col_widths(
        ws,
        {
            "A": 12.625,
            "B": 15.0,
            "C": 14.25,
            "D": 9.5,
            "E": 12.375,
            "F": 7.375,
            "G": 8.25,
            "H": 8.125,
            "I": 7.125,
            "J": 7.125,
            "K": 7.125,
            "L": 8.0,
            "M": 9.25,
            "N": 10.75,
            "O": 10.5,
            "P": 7.75,
            "Q": 7.625,
            "R": 7.75,
            "S": 8.375,
            "T": 9.5,
            "U": 21.75,
        },
    )


def export_inquiry_sheet_xlsx(
    output_path, box_meta, internal_order_no, product_details, image_paths
):
    wb = Workbook()
    ws = wb.active
    ws.title = "询价表"
    _setup_inquiry_template(ws)

    data_start = 4
    for idx, row in enumerate(box_meta, start=1):
        sku = row.get("sku") or ""
        product = product_details.get(sku, {})
        excel_row = data_start + idx - 1
        values = [
            sku,
            row.get("chinese_name") or "",
            "",
            "",
            row.get("material") or "",
            1,
            _excel_num(row.get("per_box_qty")),
            _excel_num(row.get("box_weight")),
            _excel_num(row.get("length")),
            _excel_num(row.get("width")),
            _excel_num(row.get("height")),
            None,
            None,
            None,
            None,
            _excel_num(product.get("freight_price")),
            "",
            "",
            "",
            "",
            internal_order_no if idx == 1 else "",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=excel_row, column=col, value=value)
        ws[f"M{excel_row}"] = f"=F{excel_row}*G{excel_row}"
        ws[f"N{excel_row}"] = f"=F{excel_row}*H{excel_row}"
        ws[f"O{excel_row}"] = (
            f"=I{excel_row}*J{excel_row}*K{excel_row}*F{excel_row}/1000000"
        )
        ws.row_dimensions[excel_row].height = 44

        image_path = image_paths.get(sku)
        if image_path:
            try:
                img = XLImage(image_path)
                img.width = 88
                img.height = 88
                ws.add_image(img, f"C{excel_row}")
            except Exception as exc:
                print(f"⚠️ SKU {sku} 图片写入 Excel 失败: {exc}")

    if box_meta:
        last_data_row = data_start + len(box_meta) - 1
        total_row = last_data_row + 1
        ws.append(
            [
                "合          计",
                "",
                "",
                "",
                "",
                f"=SUM(F{data_start}:F{last_data_row})",
                f"=SUM(G{data_start}:G{last_data_row})",
                f"=SUM(H{data_start}:H{last_data_row})",
                "/",
                "/",
                "/",
                "/",
                f"=SUM(M{data_start}:M{last_data_row})",
                f"=SUM(N{data_start}:N{last_data_row})",
                f"=SUM(O{data_start}:O{last_data_row})",
            ]
        )
        ws.row_dimensions[total_row].height = 42
        if last_data_row > data_start:
            ws.merge_cells(f"U{data_start}:U{last_data_row}")
        ws[f"U{data_start}"].alignment = Alignment(
            vertical="top", wrap_text=True
        )

        declare_row = total_row + 1
        ws.append(
            [
                "声明：发货人必须如实书面提供以上资料，如填写的资料与实际不符，"
                "则由此产生的一切责任及后果均由发货人承担！此委托运输的货物以公司提单为准！"
            ]
        )
        ws.merge_cells(f"A{declare_row}:U{declare_row}")
        ws.row_dimensions[declare_row].height = 42

        blank_row = declare_row + 1
        ws.append([""])
        ws.row_dimensions[blank_row].height = 9

        sign_row = blank_row + 1
        sign_values = [None] * 21
        sign_values[0] = "制单人签字："
        sign_values[6] = "发货人签字："
        sign_values[19] = "发货日期：2026年 月  日"
        ws.append(sign_values)
        ws.merge_cells(f"A{sign_row}:F{sign_row}")
        ws.merge_cells(f"G{sign_row}:M{sign_row}")
        ws.merge_cells(f"N{sign_row}:P{sign_row}")
        ws.row_dimensions[sign_row].height = 45

    wb.save(output_path)


def export_batch_excel_reports(
    order_dir,
    box_meta,
    internal_order_no,
    db_config=None,
    batch_or_order=None,
):
    if not box_meta:
        print("⚠️ 无箱位明细，跳过 Excel 导出")
        return None, None

    tag_key = (batch_or_order or "").strip() or (internal_order_no or "").strip() or "未知单号"
    inquiry_name, order_name = build_excel_report_filenames(tag_key, len(box_meta))

    print(f"\n📊 正在生成 Excel: {inquiry_name}、{order_name}...")
    product_details = fetch_product_details(
        [row.get("sku") for row in box_meta], db_config=db_config
    )
    missing = sorted(
        {
            row.get("sku")
            for row in box_meta
            if row.get("sku") and row.get("sku") not in product_details
        }
    )
    if missing:
        print(f"⚠️ 聚水潭未找到商品资料: {', '.join(missing[:8])}" + (
            f" 等共 {len(missing)} 个 SKU" if len(missing) > 8 else ""
        ))

    image_cache_dir = os.path.join(order_dir, ".img_cache")
    image_paths = prepare_product_image_paths(product_details, image_cache_dir)
    if product_details:
        print(
            f"📷 商品图片: 成功 {len(image_paths)}/{len(product_details)} 个 SKU"
        )
    inquiry_path = os.path.join(order_dir, inquiry_name)
    order_path = os.path.join(order_dir, order_name)
    export_inquiry_sheet_xlsx(
        inquiry_path,
        box_meta,
        internal_order_no or "",
        product_details,
        image_paths,
    )
    export_order_sheet_xlsx(order_path, box_meta, internal_order_no or "")
    print(f"✅ 询价表: {inquiry_path}")
    print(f"✅ 顺序表: {order_path}")
    return inquiry_path, order_path


def download_box_labels(
    supply_id,
    box_label_dir,
    headers,
    proxies,
    ordered_cargo_ids=None,
    box_meta=None,
):
    print("\n🏷️ 正在生成并下载箱唛（交货货位标签）...")
    cargo_ids = list(ordered_cargo_ids or [])
    if cargo_ids:
        print(
            f"📋 将请求 Ozon 生成 {len(cargo_ids)} 个箱唛（生成后按 ITEMS 顺序本地重排 PDF）"
        )
    else:
        try:
            cargoes = fetch_cargoes_list(supply_id, headers, proxies)
            cargo_ids = extract_cargo_ids_from_get({"cargoes": cargoes})
            if cargo_ids:
                print(
                    f"⚠️ 未拿到 create/info 顺序，cargoes/get 返回 {len(cargo_ids)} 个 cargo_id（PDF 页序可能与 ITEMS 不一致）"
                )
            else:
                print(
                    f"⚠️ cargoes/get 未返回 cargo_id（货件数 {len(cargoes)}），将仅按 supply_id 申请箱唛"
                )
        except requests.exceptions.HTTPError:
            print("⚠️ 未能通过 cargoes/get 获取 cargo_id，将仅按 supply_id 申请箱唛")

    label_create_payload = {"supply_id": int(supply_id)}
    if cargo_ids:
        label_create_payload["cargoes"] = [{"cargo_id": cid} for cid in cargo_ids]

    label_create_res = ozon_post(
        "https://api-seller.ozon.ru/v1/cargoes-label/create",
        headers,
        label_create_payload,
        proxies,
    )
    if label_create_res.get("errors"):
        print(f"❌ 箱唛生成任务创建失败: {label_create_res}")
        return None

    label_operation_id = label_create_res.get("operation_id")
    if not label_operation_id:
        print(f"❌ 未获得箱唛任务 operation_id: {label_create_res}")
        return None

    print(f"⏳ 箱唛生成任务已提交，operation_id={label_operation_id}")
    file_url = file_guid = file_content_b64 = None
    while True:
        label_get_res = ozon_post(
            "https://api-seller.ozon.ru/v1/cargoes-label/get",
            headers,
            {"operation_id": label_operation_id},
            proxies,
        )
        label_status = label_get_res.get("status", "")
        print(f"📋 箱唛生成状态: [{label_status}]")
        if label_status == "SUCCESS":
            file_url, file_guid, file_content_b64 = extract_label_file_info(
                label_get_res
            )
            break
        if label_status == "FAILED":
            print(f"❌ 箱唛生成失败: {label_get_res}")
            return None
        time.sleep(2)

    try:
        file_bytes, content_type, remote_name = download_box_label_content(
            file_url, file_guid, file_content_b64, headers, proxies
        )
    except (requests.exceptions.HTTPError, ValueError) as e:
        print(f"❌ 箱唛下载失败: {e}")
        return None

    save_name = (
        remote_name
        if remote_name
        else (
            "交货货位标签.zip"
            if "zip" in (content_type or "").lower()
            else (
                "交货货位标签.pdf"
                if "pdf" in (content_type or "").lower()
                else "交货货位标签.bin"
            )
        )
    )
    archive_path = os.path.join(box_label_dir, save_name)
    with open(archive_path, "wb") as f:
        f.write(file_bytes)
    print(f"💾 箱唛原始文件已保存: {archive_path}")

    final_path = archive_path
    is_pdf = "pdf" in (content_type or "").lower() or save_name.lower().endswith(
        ".pdf"
    )
    if cargo_ids and is_pdf and not zipfile.is_zipfile(archive_path):
        ordered_path = os.path.join(box_label_dir, "交货货位标签_按ITEMS顺序.pdf")
        try:
            ordered_bytes = reorder_cargoes_label_pdf(file_bytes, cargo_ids)
            with open(ordered_path, "wb") as f:
                f.write(ordered_bytes)
            final_path = ordered_path
            print(f"✅ 已按 ITEMS 顺序重排箱唛 PDF: {ordered_path}")
        except Exception as e:
            print(f"⚠️ 箱唛 PDF 本地重排失败: {e}")

    if zipfile.is_zipfile(archive_path):
        with zipfile.ZipFile(archive_path, "r") as zf:
            zf.extractall(box_label_dir)
        print(f"📂 已解压 zip 箱唛到: {box_label_dir}")
    return final_path


def extract_supply_id_from_order_detail(detail_res):
    supplies = extract_supplies_from_order_detail(detail_res)
    if supplies:
        return supplies[0]["supply_id"]
    return None


def extract_supplies_from_order_detail(detail_res, cluster_id_to_name=None):
    """从 order 详单解析全部 supply（多集群合并单会有多个）。"""
    cluster_id_to_name = cluster_id_to_name or {}
    normalized_id_to_name = {}
    for mc_id, name in cluster_id_to_name.items():
        if mc_id is not None:
            normalized_id_to_name[int(mc_id)] = name
    orders = detail_res.get("orders") or detail_res.get("result", {}).get(
        "orders", []
    )
    supplies_out = []
    for order in orders or []:
        for supply in order.get("supplies", []):
            supply_id = supply.get("supply_id")
            if not supply_id:
                continue
            mc_id = supply.get("macrolocal_cluster_id")
            if mc_id is not None and mc_id != "":
                mc_id = int(mc_id)
            else:
                mc_id = None
            cluster_name = normalized_id_to_name.get(mc_id, "")
            storage_wh = supply.get("storage_warehouse") or {}
            supplies_out.append(
                {
                    "supply_id": int(supply_id),
                    "macrolocal_cluster_id": mc_id,
                    "cluster_name": cluster_name,
                    "storage_warehouse_id": storage_wh.get("warehouse_id"),
                    "bundle_id": supply.get("bundle_id"),
                }
            )
    return supplies_out


def poll_order_supplies(
    headers,
    proxies,
    order_id,
    cluster_id_to_name=None,
    expected_count=1,
    max_wait_sec=45,
):
    """轮询 v3/supply-order/get，等待 supplies 列表就绪并稳定。"""
    deadline = time.time() + max_wait_sec
    last_supplies = []
    stable_rounds = 0
    while time.time() < deadline:
        detail_res = ozon_post(
            "https://api-seller.ozon.ru/v3/supply-order/get",
            headers,
            {"order_ids": [str(order_id)]},
            proxies,
        )
        supplies = extract_supplies_from_order_detail(
            detail_res, cluster_id_to_name=cluster_id_to_name
        )
        if len(supplies) >= expected_count:
            return supplies
        if supplies and supplies == last_supplies:
            stable_rounds += 1
            if stable_rounds >= 2:
                return supplies
        else:
            stable_rounds = 0
        last_supplies = supplies
        time.sleep(3)
    return last_supplies


def resolve_merged_supply_type(cluster_count, shipping_method):
    """
    多集群合并单转正必须用 MULTI_CLUSTER，才会生成每集群一个 supply。
    单集群：直发 DIRECT，中转 CROSSDOCK。
    """
    if cluster_count > 1:
        return "MULTI_CLUSTER"
    if shipping_method == "直发":
        return "DIRECT"
    return "CROSSDOCK"


def normalize_selected_cluster_warehouses(
    selected, shipping_method, cluster_count
):
    """按 supply_type 整理 selected_cluster_warehouses 字段。"""
    supply_type = resolve_merged_supply_type(cluster_count, shipping_method)
    normalized = []
    for entry in selected or []:
        mc_id = int(entry["macrolocal_cluster_id"])
        row = {"macrolocal_cluster_id": mc_id}
        if supply_type == "MULTI_CLUSTER" and shipping_method == "中转":
            bundle_id = (entry.get("bundle_id") or "").strip()
            if bundle_id:
                row["bundle_id"] = bundle_id
            elif entry.get("storage_warehouse_id"):
                row["storage_warehouse_id"] = int(entry["storage_warehouse_id"])
        elif entry.get("storage_warehouse_id"):
            row["storage_warehouse_id"] = int(entry["storage_warehouse_id"])
        elif entry.get("bundle_id"):
            row["bundle_id"] = entry["bundle_id"]
        normalized.append(row)
    return normalized


def resolve_supply_by_cluster(
    supplies, cluster_final_items, cluster_id_map, shipping_method
):
    """
    将 order 下 supplies 按 macrolocal_cluster_id 映射到集群名。
    多集群合并单需 supply_type=MULTI_CLUSTER，此处应返回与集群数一致的 supplies。
    """
    id_to_name = {int(v): k for k, v in cluster_id_map.items()}
    supply_by_cluster = {}
    for sup in supplies:
        mc_id = sup.get("macrolocal_cluster_id")
        name = sup.get("cluster_name") or ""
        if not name and mc_id is not None:
            name = id_to_name.get(int(mc_id), "")
        if name:
            supply_by_cluster[name] = sup

    expected = set(cluster_final_items.keys())
    missing = expected - set(supply_by_cluster.keys())
    if missing and shipping_method == "中转" and len(supplies) == 1:
        print(
            "❌ 中转多集群仅得到 1 个 supply，说明转正时 supply_type 应为 "
            "MULTI_CLUSTER 而非 CROSSDOCK"
        )
    return supply_by_cluster


def split_cargo_ids_by_cluster(shipment_items, ordered_cargo_ids):
    """将单 supply 返回的 cargo_id 序列按集群拆分（与 shipment_items 箱序一致）。"""
    shipment_items = sort_shipment_items_by_row_id(shipment_items)
    by_cluster = defaultdict(list)
    idx = 0
    for row in shipment_items:
        cluster = row.get("cluster") or "未知集群"
        per_box = int(row["number"])
        qty = int(row["quantity"])
        box_count = qty // per_box if per_box > 0 else 0
        for _ in range(box_count):
            if idx >= len(ordered_cargo_ids):
                raise ValueError(
                    f"集群 {cluster!r} 拆分 cargo_id 时索引越界 "
                    f"({idx}/{len(ordered_cargo_ids)})"
                )
            by_cluster[cluster].append(int(ordered_cargo_ids[idx]))
            idx += 1
    if idx != len(ordered_cargo_ids):
        raise ValueError(
            f"cargo_id 数量 {len(ordered_cargo_ids)} 与箱数 {idx} 不一致"
        )
    return dict(by_cluster)


def build_shop_session(target_shop):
    if target_shop not in SHOP_DATA:
        raise KeyError(f"未找到店铺【{target_shop}】的 API 配置")
    client_id = SHOP_DATA[target_shop]["client_id"]
    api_key = SHOP_DATA[target_shop]["api_key"]
    proxies = {"http": None, "https": None}
    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }
    return headers, proxies


def prepare_shipment_items(items, headers, proxies):
    """将数据库 ITEMS 转为 shipment_items 与草稿用 final_items。"""
    offer_ids = list(dict.fromkeys(item["sku"] for item in items))
    sku_data = ozon_post(
        "https://api-seller.ozon.ru/v3/product/info/list",
        headers,
        {"offer_id": offer_ids},
        proxies,
    )

    offer_to_sku_map = {}
    offer_to_barcode_map = {}
    for p in sku_data.get("items", []):
        real_sku = p.get("sku")
        oid = p.get("offer_id")
        offer_to_sku_map[oid] = real_sku
        barcodes = p.get("barcodes") or []
        if barcodes:
            offer_to_barcode_map[oid] = barcodes[0]

    shipment_items = []
    draft_qty_by_sku = {}
    for item in items:
        user_offer_id = item["sku"]
        real_numeric_sku = offer_to_sku_map.get(user_offer_id)
        if not real_numeric_sku:
            print(
                f"⚠️ 警告: 货号 '{user_offer_id}' 在后台未查到任何有效的数字 SKU！该商品将被跳过。"
            )
            continue
        sku_int = int(real_numeric_sku)
        draft_qty_by_sku[sku_int] = (
            draft_qty_by_sku.get(sku_int, 0) + item["quantity"]
        )
        shipment_items.append(
            {
                "offer_id": user_offer_id,
                "quantity": item["quantity"],
                "number": item["number"],
                "barcode": offer_to_barcode_map.get(user_offer_id),
                "cluster": item.get("cluster") or "",
                "row_id": item.get("row_id"),
                **item_meta_kwargs(item),
            }
        )

    final_items = [
        {"sku": sku, "quantity": qty} for sku, qty in draft_qty_by_sku.items()
    ]
    total_boxes = sum(
        row["quantity"] // row["number"]
        for row in shipment_items
        if row["number"] > 0 and row["quantity"] % row["number"] == 0
    )
    return shipment_items, final_items, total_boxes


def prepare_merged_shipment(group, headers, proxies):
    """
    多集群合并：解析全部 ITEMS，按集群分别汇总草稿 SKU 数量。
    返回 (shipment_items, cluster_final_items, cluster_id_map, total_boxes)。
    """
    items = group["items"]
    offer_ids = list(dict.fromkeys(item["sku"] for item in items))
    sku_data = ozon_post(
        "https://api-seller.ozon.ru/v3/product/info/list",
        headers,
        {"offer_id": offer_ids},
        proxies,
    )

    offer_to_sku_map = {}
    offer_to_barcode_map = {}
    for p in sku_data.get("items", []):
        real_sku = p.get("sku")
        oid = p.get("offer_id")
        offer_to_sku_map[oid] = real_sku
        barcodes = p.get("barcodes") or []
        if barcodes:
            offer_to_barcode_map[oid] = barcodes[0]

    shipment_items = []
    cluster_final_items = {}
    cluster_id_map = {}
    cluster_draft_qty = defaultdict(lambda: defaultdict(int))

    for item in items:
        cluster_name = (item.get("cluster") or "").strip() or "未知集群"
        user_offer_id = item["sku"]
        real_numeric_sku = offer_to_sku_map.get(user_offer_id)
        if not real_numeric_sku:
            print(
                f"⚠️ 警告: 集群 {cluster_name!r} 货号 '{user_offer_id}' "
                f"未查到数字 SKU，跳过。"
            )
            continue
        sku_int = int(real_numeric_sku)
        cluster_draft_qty[cluster_name][sku_int] += item["quantity"]
        shipment_items.append(
            {
                "offer_id": user_offer_id,
                "quantity": item["quantity"],
                "number": item["number"],
                "barcode": offer_to_barcode_map.get(user_offer_id),
                "cluster": cluster_name,
                "row_id": item.get("row_id"),
                **item_meta_kwargs(item),
            }
        )

    for cluster_name, qty_map in cluster_draft_qty.items():
        cluster_final_items[cluster_name] = [
            {"sku": sku, "quantity": qty} for sku, qty in qty_map.items()
        ]
        try:
            cluster_id_map[cluster_name] = resolve_macrolocal_cluster_id(
                cluster_name, headers, proxies
            )
        except ValueError as e:
            raise ValueError(f"集群 {cluster_name!r}: {e}") from e

    if not shipment_items or not cluster_final_items:
        skipped = [
            item["sku"]
            for item in items
            if item["sku"] not in offer_to_sku_map
        ]
        hint = (
            f"以下货号未在 Ozon 查到 SKU: {', '.join(skipped)}"
            if skipped
            else "货号 SKU 映射后无有效发货行"
        )
        raise ValueError(f"无有效 ITEMS（{hint}）")

    shipment_items = sort_shipment_items_by_row_id(shipment_items)
    total_boxes = sum(
        row["quantity"] // row["number"]
        for row in shipment_items
        if row["number"] > 0 and row["quantity"] % row["number"] == 0
    )
    return shipment_items, cluster_final_items, cluster_id_map, total_boxes


def extract_cargoes_create_error_reasons(info_res):
    errors = info_res.get("errors") or {}
    if isinstance(errors, dict):
        reasons = errors.get("error_reasons") or []
        if reasons:
            return list(reasons)
    result = info_res.get("result") or {}
    if isinstance(result, dict):
        reasons = result.get("error_reasons") or []
        if reasons:
            return list(reasons)
    return []


def shipment_item_row_ids(items):
    """从 shipment_items 提取数据库唯一ID（字符串，去重保序）。"""
    ids = []
    for row in items or []:
        rid = row.get("row_id")
        if rid is None or rid == "":
            continue
        ids.append(str(int(rid)))
    return list(dict.fromkeys(ids))


def shipment_item_skus(items):
    """从 shipment_items 提取 offer_id/SKU（去重保序）。"""
    skus = []
    for row in items or []:
        sku = (row.get("offer_id") or row.get("sku") or "").strip()
        if sku and sku not in skus:
            skus.append(sku)
    return skus


def format_cluster_cargo_failure(
    cluster_name, cluster_items, cargo_err=None, max_items=20
):
    """
    货位提交失败时拼可读明细：原因 + 集群 + SKU + 唯一ID。
    Ozon 常只返回 VALIDATION_FAILED，用本集群提交清单反查登记行。
    """
    reason_text = (cargo_err or "").strip() or "货位提交失败(未知原因)"
    skus = shipment_item_skus(cluster_items)
    ids = shipment_item_row_ids(cluster_items)

    def _join(values):
        if not values:
            return "-"
        shown = values[:max_items]
        text = ",".join(shown)
        if len(values) > max_items:
            text += f"等{len(values)}个"
        return text

    return (
        f"{reason_text} 集群={cluster_name} "
        f"SKU={_join(skus)} 唯一ID={_join(ids)}"
    )


def build_cargo_failure_meta(failed_cluster, failed_items, ok_items_by_cluster):
    """组装 runner 用的部分失败 meta。"""
    failed_row_ids = shipment_item_row_ids(failed_items)
    ok_row_ids = []
    for items in (ok_items_by_cluster or {}).values():
        ok_row_ids.extend(shipment_item_row_ids(items))
    ok_row_ids = list(dict.fromkeys(ok_row_ids))
    return {
        "failed_cluster": failed_cluster,
        "failed_skus": shipment_item_skus(failed_items),
        "failed_row_ids": failed_row_ids,
        "ok_row_ids": ok_row_ids,
    }


def print_warehouse_limits_exceed_hint(
    supply_id, order_id, submitted_count, total_count, shop=None, batch_no=None
):
    shop = shop or "YOUR_SHOP"
    batch_no = batch_no or "YOUR_BATCH"
    print("\n" + "=" * 60)
    print("⚠️ 供货单总箱数超过当前上限（WAREHOUSE_LIMITS_EXCEED）")
    print("   可能未成功执行 transport/activate，或 activate 在已有货位之后调用。")
    print(f"   order_id={order_id}  supply_id={supply_id}")
    print(f"   已写入 {submitted_count}/{total_count} 箱。")
    print("\n   若 supply 尚无货位，可尝试断点续传（会先 activate 再补货位）：")
    print(
        "   python Ozon\\全流程.py --resume-cargoes "
        f"--shop {shop} --batch {batch_no} --order-id {order_id} "
        f"--supply-id {supply_id}"
    )
    print("=" * 60 + "\n")


def activate_cargoes_transport(supply_id, headers, proxies):
    """
    激活运输货位包（API 等效于后台「提升箱子上限的运输货位」）。
    必须在第一次 cargoes/create 之前调用。
    """
    print("\n🚛 [STEP 6.5] 正在激活运输货位包（提升箱子上限）...")
    activate_res = ozon_post(
        "https://api-seller.ozon.ru/v1/cargoes/transport/activate",
        headers,
        {"supply_id": int(supply_id), "is_transport": True},
        proxies,
    )
    operation_id = activate_res.get("operation_id")
    if not operation_id:
        print(f"❌ transport/activate 未返回 operation_id: {activate_res}")
        return False

    print(f"⏳ 运输货位激活任务 operation_id={operation_id}")
    while True:
        status_res = ozon_post(
            "https://api-seller.ozon.ru/v1/cargoes/transport/activate/status",
            headers,
            {"operation_id": operation_id},
            proxies,
        )
        activate_status = status_res.get("status", "")
        print(f"📋 运输货位激活状态: [{activate_status}]")
        if activate_status == "SUCCESS":
            print("✅ 运输货位包已激活，可提交超过 30 箱的货位")
            return True
        if activate_status == "FAILED":
            reasons = status_res.get("error_reasons") or []
            print(f"❌ 运输货位激活失败: {status_res}")
            if "CAN_NOT_EDIT_TAG" in reasons:
                print(
                    "   该 supply 已有货位，无法再激活。"
                    "请取消供货单后重建，或使用空货位的新 supply。"
                )
            return False
        time.sleep(2)


def ensure_transport_activated_for_large_supply(
    supply_id, total_boxes, headers, proxies, existing_box_count=0
):
    """箱数 >30 且尚无盒子货位时，先 activate。"""
    if total_boxes <= MAX_BOXES_PER_SUPPLY_ORDER:
        return True
    if existing_box_count > 0:
        print(
            f"⚠️ 本 supply 已有 {existing_box_count} 个盒子货位且总箱数 {total_boxes} > "
            f"{MAX_BOXES_PER_SUPPLY_ORDER}，跳过 activate（可能已激活或无法补救）"
        )
        return True
    print(
        f"\n📦 本单共 {total_boxes} 箱（>{MAX_BOXES_PER_SUPPLY_ORDER}），"
        f"将通过 API 激活运输货位包并创建运输货位"
    )
    return activate_cargoes_transport(supply_id, headers, proxies)


def fetch_transport_cargo_ids(supply_id, headers, proxies):
    """获取 supply 下已创建的运输货位（API 类型 PALLET）。"""
    cargoes = fetch_cargoes_list(supply_id, headers, proxies)
    ids = []
    for cargo in cargoes:
        if (cargo.get("type") or "").upper() != "PALLET":
            continue
        cargo_id = cargo.get("cargo_id") or cargo.get("id")
        if cargo_id:
            ids.append(int(cargo_id))
    return ids


def count_box_cargoes(supply_id, headers, proxies):
    cargoes = fetch_cargoes_list(supply_id, headers, proxies)
    return sum(
        1 for c in cargoes if (c.get("type") or "").upper() == "BOX"
    )


def create_transport_cargo_slot(
    supply_id, total_boxes, headers, proxies, transport_count=1
):
    """
    创建运输货位（网页：添加货位 → 盒子=总箱数，运输货位=1）。
    API: type=PALLET, boxes_count=总箱数, count=运输货位数。
    """
    print(
        f"\n🚚 [STEP 6.6] 正在创建运输货位"
        f"（盒子={total_boxes}，运输货位={transport_count}）..."
    )
    create_res = ozon_post(
        "https://api-seller.ozon.ru/v1/cargoes/transport/create",
        headers,
        {
            "supply_id": int(supply_id),
            "transport_cargoes": [
                {
                    "type": "PALLET",
                    "boxes_count": int(total_boxes),
                    "count": int(transport_count),
                }
            ],
        },
        proxies,
    )
    if create_res.get("error_reasons"):
        print(f"❌ transport/create 返回错误: {create_res}")
        return None

    operation_id = create_res.get("operation_id")
    if not operation_id:
        print(f"❌ transport/create 未返回 operation_id: {create_res}")
        return None

    print(f"⏳ 运输货位创建任务 operation_id={operation_id}")
    while True:
        status_res = ozon_post(
            "https://api-seller.ozon.ru/v1/cargoes/transport/create/status",
            headers,
            {"operation_id": operation_id},
            proxies,
        )
        create_status = status_res.get("status", "")
        print(f"📋 运输货位创建状态: [{create_status}]")
        if create_status == "SUCCESS":
            transport_cargoes = (status_res.get("result") or {}).get(
                "transport_cargoes"
            ) or []
            if transport_cargoes:
                transport_id = transport_cargoes[0].get("id")
                print(f"✅ 运输货位已创建 transport_cargo_id={transport_id}")
                return int(transport_id)
            print(f"❌ 未解析到 transport_cargo_id: {status_res}")
            return None
        if create_status == "FAILED":
            print(f"❌ 运输货位创建失败: {status_res}")
            return None
        time.sleep(2)


def bind_boxes_to_transport_cargo(
    supply_id, transport_cargo_id, box_cargo_ids, headers, proxies
):
    """将盒子货位绑定到运输货位（对应网页分配箱子到运输货位）。"""
    if not box_cargo_ids:
        print("⚠️ 无盒子货位可绑定，跳过 transport/bind")
        return True

    print(
        f"\n🔗 [STEP 7.5] 正在将 {len(box_cargo_ids)} 个盒子"
        f"绑定到运输货位 {transport_cargo_id}..."
    )
    bind_batch_size = 1000
    for start in range(0, len(box_cargo_ids), bind_batch_size):
        chunk = box_cargo_ids[start : start + bind_batch_size]
        bind_res = ozon_post(
            "https://api-seller.ozon.ru/v1/cargoes/transport/bind",
            headers,
            {
                "supply_id": int(supply_id),
                "transport_cargo_bind": [
                    {
                        "transport_cargo_id": int(transport_cargo_id),
                        "cargo_ids": chunk,
                    }
                ],
            },
            proxies,
        )
        if bind_res.get("error_reasons"):
            print(f"❌ transport/bind 返回错误: {bind_res}")
            return False

        operation_id = bind_res.get("operation_id")
        if not operation_id:
            print(f"❌ transport/bind 未返回 operation_id: {bind_res}")
            return False

        while True:
            status_res = ozon_post(
                "https://api-seller.ozon.ru/v1/cargoes/transport/bind/status",
                headers,
                {"operation_id": operation_id},
                proxies,
            )
            bind_status = status_res.get("status", "")
            print(f"📋 运输货位绑定状态: [{bind_status}]")
            if bind_status == "SUCCESS":
                break
            if bind_status == "FAILED":
                print(f"❌ 运输货位绑定失败: {status_res}")
                return False
            time.sleep(2)

    print("✅ 盒子已全部绑定到运输货位")
    return True


def ensure_large_supply_transport_setup(
    supply_id, total_boxes, headers, proxies, existing_box_count=0
):
    """
    >30 箱时：activate → 创建 1 个运输货位（boxes_count=总箱数）。
    返回 transport_cargo_id；≤30 箱返回 None。
    """
    if total_boxes <= MAX_BOXES_PER_SUPPLY_ORDER:
        return None

    existing_transport_ids = fetch_transport_cargo_ids(
        supply_id, headers, proxies
    )
    if existing_transport_ids:
        print(
            f"✅ 已存在运输货位 transport_cargo_id={existing_transport_ids[0]}"
        )
        return existing_transport_ids[0]

    if existing_box_count > 0:
        print(
            f"⚠️ 已有 {existing_box_count} 个盒子货位但未找到运输货位，"
            f"无法自动创建（需空货位 supply 重建）"
        )
        return None

    if not ensure_transport_activated_for_large_supply(
        supply_id, total_boxes, headers, proxies, existing_box_count=0
    ):
        return None
    time.sleep(1.5)

    return create_transport_cargo_slot(
        supply_id, total_boxes, headers, proxies, transport_count=1
    )


def fetch_ordered_box_cargo_ids(supply_id, cargoes_list, headers, proxies):
    ordered = fetch_ordered_cargo_ids_by_keys(
        supply_id, cargoes_list, headers, proxies
    )
    if len(ordered) == len(cargoes_list):
        return ordered

    cargoes = fetch_cargoes_list(supply_id, headers, proxies)
    box_ids = [
        int(c["cargo_id"])
        for c in cargoes
        if (c.get("type") or "").upper() == "BOX" and c.get("cargo_id")
    ]
    if len(box_ids) >= len(cargoes_list):
        return box_ids[: len(cargoes_list)]
    return ordered


def fetch_ordered_cargo_ids_by_keys(supply_id, cargoes_list, headers, proxies):
    """按 box-xxx 键顺序从 supply 中解析 cargo_id。"""
    cargoes = fetch_cargoes_list(supply_id, headers, proxies)
    key_to_id = {}
    for cargo in cargoes:
        key = cargo.get("key") or cargo.get("cargo_key")
        cargo_id = cargo.get("cargo_id") or cargo.get("id")
        if key and cargo_id:
            key_to_id[key] = int(cargo_id)

    ordered = []
    for cargo in cargoes_list:
        key = cargo.get("key")
        if key and key in key_to_id:
            ordered.append(key_to_id[key])

    if len(ordered) == len(cargoes_list):
        return ordered

    fallback = [
        int(c["cargo_id"])
        for c in cargoes
        if (c.get("type") or "").upper() == "BOX" and c.get("cargo_id")
    ]
    if fallback:
        print(
            f"⚠️ 未能按 box 键完整匹配 cargo_id（{len(ordered)}/{len(cargoes_list)}），"
            f"改用 cargoes/get 中 BOX 顺序"
        )
        return fallback[: len(cargoes_list)]
    return ordered


def poll_cargoes_create_status(
    operation_id, supply_id, headers, proxies, expected_box_count
):
    status_urls = [
        "https://api-seller.ozon.ru/v2/cargoes/create/info",
        "https://api-seller.ozon.ru/v1/cargoes/create/info",
    ]
    for status_url in status_urls:
        print(f"⏳ 正在通过 {status_url} 查询货件写入进度...")
        empty_success_rounds = 0
        while True:
            try:
                info_res = ozon_post(
                    status_url,
                    headers,
                    {"operation_id": operation_id},
                    proxies,
                )
            except requests.exceptions.HTTPError as e:
                if e.response is not None and e.response.status_code == 404:
                    break
                raise

            cargo_status = info_res.get("status", "")
            print(f"📋 货件写入状态: [{cargo_status}]")
            if cargo_status == "SUCCESS":
                ordered_ids = extract_ordered_cargo_ids_from_create_info(
                    info_res
                )
                if ordered_ids:
                    return True, ordered_ids, []
                empty_success_rounds += 1
                if empty_success_rounds >= 15:
                    return True, [], []
                print(
                    "⏳ 货件写入 SUCCESS 但 create/info 尚未返回 cargo_id，继续轮询..."
                )
                time.sleep(2)
                continue
            if cargo_status == "FAILED":
                errors = info_res.get("errors") or info_res.get("result")
                reasons = extract_cargoes_create_error_reasons(info_res)
                print(f"❌ 货件写入失败: {errors or info_res}")
                return False, [], reasons
            time.sleep(2)

    try:
        cargoes = fetch_cargoes_list(supply_id, headers, proxies)
        actual_count = sum(
            1 for c in cargoes if (c.get("type") or "").upper() == "BOX"
        )
        if actual_count >= expected_box_count:
            return True, [], []
    except Exception:
        pass
    return False, [], []


def submit_cargoes_in_batches(
    supply_id,
    cargoes_list,
    headers,
    proxies,
    batch_size=CARGOES_CREATE_BATCH_SIZE,
    append_only=False,
    order_id=None,
    batch_no=None,
    shop=None,
    already_submitted=0,
    full_total=None,
):
    """分批提交货位（单次最多 batch_size 箱，后续批次追加不覆盖）。"""
    total = len(cargoes_list)
    if total == 0:
        raise ValueError("货位列表为空")

    batch_count = (total + batch_size - 1) // batch_size
    if batch_count > 1:
        print(
            f"📦 共 {total} 箱，将分 {batch_count} 批提交（每批最多 {batch_size} 箱）"
        )

    all_ordered_ids = []
    submitted_so_far = 0
    for batch_idx in range(batch_count):
        chunk = cargoes_list[
            batch_idx * batch_size : (batch_idx + 1) * batch_size
        ]
        print(
            f"📦 提交货位批次 [{batch_idx + 1}/{batch_count}]，本批 {len(chunk)} 箱..."
        )
        delete_current = (batch_idx == 0) and not append_only
        cargoes_res = ozon_post(
            "https://api-seller.ozon.ru/v1/cargoes/create",
            headers,
            {
                "supply_id": supply_id,
                "delete_current_version": delete_current,
                "cargoes": chunk,
            },
            proxies,
        )
        operation_id = cargoes_res.get("operation_id")
        if not operation_id:
            err = f"货位创建未返回 operation_id: {_api_err_snippet(cargoes_res)}"
            print(f"❌ {err}")
            return False, all_ordered_ids, err

        ok, ordered_ids, reasons = poll_cargoes_create_status(
            operation_id, supply_id, headers, proxies, len(chunk)
        )
        if not ok:
            if "WAREHOUSE_LIMITS_EXCEED" in reasons and order_id:
                print_warehouse_limits_exceed_hint(
                    supply_id,
                    order_id,
                    already_submitted + submitted_so_far,
                    full_total or (already_submitted + total),
                    shop=shop,
                    batch_no=batch_no,
                )
            reason_text = ", ".join(reasons) if reasons else "未知原因"
            err = f"货位提交失败({reason_text})"
            return False, all_ordered_ids, err
        if not ordered_ids:
            ordered_ids = fetch_new_box_cargo_ids(
                supply_id, all_ordered_ids, len(chunk), headers, proxies
            )
            if len(ordered_ids) < len(chunk):
                print(
                    f"⚠️ 本批仅解析到 {len(ordered_ids)}/{len(chunk)} 个 cargo_id"
                )
        all_ordered_ids.extend(ordered_ids)
        submitted_so_far += len(chunk)
        if batch_idx < batch_count - 1:
            time.sleep(1.5)

    return True, all_ordered_ids, ""


def build_cargoes_from_items(shipment_items):
    cargoes = []
    box_idx = 1
    for row_idx, row in enumerate(shipment_items, start=1):
        offer_id = row["offer_id"]
        total_qty = row["quantity"]
        per_box = row["number"]
        if per_box <= 0 or total_qty % per_box != 0:
            raise ValueError(
                f"行 {row_idx} 货号 {offer_id} 数量配置错误，无法被整除。"
            )
        box_count = total_qty // per_box
        item_payload = {"offer_id": offer_id, "quant": 1, "quantity": per_box}
        if row.get("barcode"):
            item_payload["barcode"] = row["barcode"]

        for _ in range(box_count):
            cargoes.append(
                {
                    "key": f"box-{box_idx:03d}",
                    "value": {"type": "BOX", "items": [dict(item_payload)]},
                }
            )
            box_idx += 1
    return cargoes


# ==================== 中转（CROSSDOCK）特有核心逻辑 ====================
def resolve_macrolocal_cluster_id(cluster_name, headers, proxies):
    url = "https://api-seller.ozon.ru/v1/cluster/list"
    target = cluster_name.strip()
    matched_ids = []
    for cluster_type in ("CLUSTER_TYPE_OZON", "CLUSTER_TYPE_CIS"):
        data = ozon_post(url, headers, {"cluster_type": cluster_type}, proxies)
        for cluster in data.get("clusters", []):
            if (cluster.get("name") or "").strip() == target:
                mc_id = cluster.get("macrolocal_cluster_id")
                if mc_id is not None:
                    matched_ids.append(int(mc_id))

    if not matched_ids:
        raise ValueError(
            f"未在 Ozon 集群列表中找到名称 {cluster_name!r}，请核对 MACROLOCAL_CLUSTER。"
        )
    unique_ids = list(dict.fromkeys(matched_ids))
    return unique_ids[0]


def normalize_drop_off_warehouse_type(raw_type):
    if not raw_type:
        return "SORTING_CENTER"
    key = str(raw_type).upper()
    mapping = {
        "WAREHOUSE_TYPE_DELIVERY_POINT": "DELIVERY_POINT",
        "WAREHOUSE_TYPE_ORDERS_RECEIVING_POINT": "ORDERS_RECEIVING_POINT",
        "WAREHOUSE_TYPE_SORTING_CENTER": "SORTING_CENTER",
        "WAREHOUSE_TYPE_FULL_FILLMENT": "FULL_FILLMENT",
        "WAREHOUSE_TYPE_CROSS_DOCK": "CROSS_DOCK",
    }
    return mapping.get(key, key.replace("WAREHOUSE_TYPE_", ""))


def haversine_km(lat1, lon1, lat2, lon2):
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(p1) * math.cos(p2) * math.sin(dlon / 2) ** 2
    )
    return 2 * r * math.asin(math.sqrt(a))


def fbo_search_keywords_for_name(warehouse_name):
    keywords = {"москва", "МСК_"}
    name = (warehouse_name or "").strip()
    if not name:
        return keywords
    for part in re.split(r"[_\-\s]+", name):
        if len(part) >= 4:
            keywords.add(part)
    if "СТРОG" in name.upper():
        keywords.add("строг")
    return keywords


def fetch_fbo_crossdock_dropoffs(headers, proxies, search_keywords):
    rows = []
    seen = set()
    for keyword in search_keywords:
        if len(keyword) < 4:
            continue
        try:
            res = ozon_post(
                "https://api-seller.ozon.ru/v1/warehouse/fbo/list",
                headers,
                {
                    "search": keyword,
                    "filter_by_supply_type": ["CREATE_TYPE_CROSSDOCK"],
                },
                proxies,
            )
            for item in res.get("search") or []:
                wh_id = item.get("warehouse_id")
                if not wh_id or int(wh_id) in seen:
                    continue
                seen.add(int(wh_id))
                coord = item.get("coordinates") or {}
                rows.append(
                    {
                        "warehouse_id": int(wh_id),
                        "name": (item.get("name") or "").strip(),
                        "warehouse_type": normalize_drop_off_warehouse_type(
                            item.get("warehouse_type")
                        ),
                        "latitude": coord.get("latitude"),
                        "longitude": coord.get("longitude"),
                        "address": item.get("address") or "",
                        "source": "warehouse_fbo_list",
                    }
                )
        except Exception:
            continue
    return rows


def fetch_cluster_dropoff_candidates(headers, proxies, macrolocal_cluster_id):
    dropoff_types = {t.upper() for t in CROSSDOCK_DROP_OFF_WAREHOUSE_TYPES}
    candidates = {}
    for cluster_type in ("CLUSTER_TYPE_OZON", "CLUSTER_TYPE_CIS"):
        data = ozon_post(
            "https://api-seller.ozon.ru/v1/cluster/list",
            headers,
            {"cluster_type": cluster_type},
            proxies,
        )
        for cluster in data.get("clusters") or []:
            if cluster.get("macrolocal_cluster_id") != macrolocal_cluster_id:
                continue
            for lc in cluster.get("logistic_clusters") or []:
                for wh in lc.get("warehouses") or []:
                    if (wh.get("type") or "").upper() not in dropoff_types:
                        continue
                    wh_id = int(wh["warehouse_id"])
                    candidates[wh_id] = {
                        "warehouse_id": wh_id,
                        "name": (wh.get("name") or "").strip(),
                        "warehouse_type": normalize_drop_off_warehouse_type(
                            wh.get("type")
                        ),
                        "latitude": None,
                        "longitude": None,
                        "address": "",
                        "source": "cluster_list",
                    }
    return list(candidates.values())


def fetch_all_clusters_dropoff_candidates(headers, proxies, cluster_id_map):
    """多集群中转：合并全部集群在 cluster/list 中的交接仓候选。"""
    merged = {}
    for mc_id in cluster_id_map.values():
        for row in fetch_cluster_dropoff_candidates(headers, proxies, mc_id):
            merged[row["warehouse_id"]] = row
    return list(merged.values())


def is_multi_cluster_client_payload_error(response_text):
    """400 且为请求体结构/items 问题时不应换交接仓重试。"""
    text = response_text or ""
    return "ClustersInfo" in text and "Items" in text


def merge_dropoff_candidates(*sources):
    merged = {}
    for rows in sources:
        for row in rows:
            wid = row["warehouse_id"]
            if wid not in merged:
                merged[wid] = dict(row)
                continue
            for key, val in row.items():
                if val not in (None, "", []):
                    merged[wid][key] = val
    return list(merged.values())


def build_dropoff_try_order(
    candidates, preferred_name, max_nearby_km=80, max_fallback=15
):
    target = preferred_name.strip()
    preferred = next((c for c in candidates if c["name"] == target), None)
    if not preferred:
        raise ValueError(f"未找到中转站 {target!r}，请核对名字。")

    pref_lat, pref_lon = preferred.get("latitude"), preferred.get("longitude")
    nearby = []
    for c in candidates:
        if c["warehouse_id"] == preferred["warehouse_id"]:
            continue
        name_upper = (c.get("name") or "").upper()
        if "ХАБ" not in name_upper:
            continue
        lat, lon = c.get("latitude"), c.get("longitude")
        if lat is None or lon is None or pref_lat is None or pref_lon is None:
            continue
        dist = haversine_km(pref_lat, pref_lon, lat, lon)
        if dist <= max_nearby_km:
            item = dict(c)
            item["distance_km"] = round(dist, 1)
            nearby.append(item)

    random.shuffle(nearby)
    nearby.sort(
        key=lambda x: (
            x.get("distance_km") or 9999,
        )
    )
    return [preferred] + nearby[:max_fallback]


def create_crossdock_draft_and_poll_storage(
    dropoff_wh,
    final_items,
    macrolocal_cluster_id,
    delivery_type,
    seller_warehouse_id,
    headers,
    proxies,
):
    delivery_info = (
        {"type": "PICKUP", "seller_warehouse_id": int(seller_warehouse_id)}
        if delivery_type.upper() == "PICKUP"
        else {
            "type": "DROPOFF",
            "drop_off_warehouse": {
                "warehouse_id": int(dropoff_wh["warehouse_id"]),
                "warehouse_type": dropoff_wh["warehouse_type"],
            },
        }
    )

    draft_payload = {
        "cluster_info": {
            "macrolocal_cluster_id": macrolocal_cluster_id,
            "items": final_items,
        },
        "delivery_info": delivery_info,
        "deletion_sku_mode": "PARTIAL",
    }
    try:
        draft_res = ozon_post(
            "https://api-seller.ozon.ru/v1/draft/crossdock/create",
            headers,
            draft_payload,
            proxies,
        )
    except Exception:
        return None, None, [], "CREATE_FAILED"

    draft_id = draft_res.get("draft_id")
    errors = draft_res.get("errors") or []
    if not draft_id:
        return None, None, errors, "CREATE_FAILED"

    time.sleep(2)
    while True:
        info_res = ozon_post(
            "https://api-seller.ozon.ru/v2/draft/create/info",
            headers,
            {"draft_id": draft_id},
            proxies,
        )
        current_status = info_res.get("status")
        print(f"📋 当前中转任务状态: [{current_status}]")

        if current_status == "SUCCESS":
            for cluster in extract_clusters_from_draft_info(info_res):
                entry, _, _ = build_cluster_warehouse_entry(
                    cluster, macrolocal_cluster_id, crossdock=True
                )
                if entry and (
                    entry.get("storage_warehouse_id") or entry.get("bundle_id")
                ):
                    warehouse_extra = {
                        k: v
                        for k, v in entry.items()
                        if k != "macrolocal_cluster_id"
                    }
                    return draft_id, warehouse_extra, errors, "SUCCESS"
            return draft_id, {}, errors, "SUCCESS"
        if current_status == "FAILED":
            return draft_id, None, info_res.get("errors") or [], "FAILED"
        if current_status in ("IN_PROGRESS", "UNSPECIFIED"):
            time.sleep(3)
            continue
        return draft_id, None, errors, "UNKNOWN"


GOOD_WAREHOUSE_STATE_HINTS = (
    "FULL_AVAILABLE",
    "PARTIAL_AVAILABLE",
)


def _warehouse_availability_state(wh):
    avail = wh.get("availability_status") or {}
    status = wh.get("status") or {}
    return (avail.get("state") or status.get("state") or "").strip()


def _warehouse_state_rank(state):
    upper = (state or "").upper()
    if "FULL_AVAILABLE" in upper:
        return 0
    if "PARTIAL_AVAILABLE" in upper:
        return 1
    return 99


def _warehouse_is_available(wh):
    status = wh.get("status") or {}
    if status.get("is_available") is True:
        return True
    state = _warehouse_availability_state(wh)
    if not state:
        return False
    upper = state.upper()
    return any(hint in upper for hint in GOOD_WAREHOUSE_STATE_HINTS)


def _warehouse_storage_id(wh):
    for key in ("storage_warehouse", "supply_warehouse"):
        nested = wh.get(key) or {}
        wid = nested.get("warehouse_id")
        if wid:
            return int(wid)
    wid = wh.get("warehouse_id")
    if wid:
        return int(wid)
    return None


def _warehouse_storage_name(wh):
    for key in ("storage_warehouse", "supply_warehouse"):
        nested = wh.get(key) or {}
        name = (nested.get("name") or "").strip()
        if name:
            return name
    return (wh.get("name") or "").strip()


def extract_clusters_from_draft_info(info_res):
    clusters = info_res.get("clusters")
    if clusters:
        return clusters
    result = info_res.get("result") or {}
    return result.get("clusters") or []


def _warehouse_invalid_reason(wh):
    avail = wh.get("availability_status") or {}
    status = wh.get("status") or {}
    return (
        avail.get("invalid_reason")
        or status.get("invalid_reason")
        or ""
    ).strip()


def is_sku_matrix_unavailable(cluster_block):
    """Ozon: NOT_AVAILABLE_MATRIX = 该 SKU 不在集群可发矩阵（换交接仓无效）。"""
    for wh in cluster_block.get("warehouses") or []:
        reason = _warehouse_invalid_reason(wh).upper()
        if "NOT_AVAILABLE_MATRIX" in reason:
            return True
    return False


def summarize_cluster_storage_rejection(cluster_block, cluster_label, skus=None):
    """把算仓 SUCCESS 但无可用仓，拼成可读原因（含矩阵拒绝）。"""
    warehouses = cluster_block.get("warehouses") or []
    sku_hint = ""
    if skus:
        shown = [str(s) for s in skus[:8]]
        sku_hint = f" SKU={','.join(shown)}"
        if len(skus) > 8:
            sku_hint += f"等{len(skus)}个"

    if not warehouses:
        return f"集群 {cluster_label} 无候选仓{sku_hint}"

    parts = []
    for wh in warehouses[:5]:
        state = _warehouse_availability_state(wh) or "-"
        reason = _warehouse_invalid_reason(wh) or "-"
        wh_name = _warehouse_storage_name(wh) or "(无名)"
        parts.append(f"{wh_name} state={state} reason={reason}")

    detail = "; ".join(parts)
    if is_sku_matrix_unavailable(cluster_block):
        return (
            f"集群 {cluster_label} 不能接受该SKU"
            f"（NOT_AVAILABLE_MATRIX，不在可发矩阵）{sku_hint}"
            f" [{detail}]"
        )
    return f"集群 {cluster_label} 无可用仓{sku_hint} [{detail}]"


def log_cluster_warehouse_diagnostics(cluster_block, cluster_label):
    warehouses = cluster_block.get("warehouses") or []
    print(f"⚠️ 集群 {cluster_label} 算仓明细（{len(warehouses)} 个候选仓）:")
    for idx, wh in enumerate(warehouses[:8], start=1):
        state = _warehouse_availability_state(wh)
        reason = _warehouse_invalid_reason(wh)
        wh_id = _warehouse_storage_id(wh)
        wh_name = _warehouse_storage_name(wh)
        bundle_id = (wh.get("bundle_id") or "").strip()
        is_avail = _warehouse_is_available(wh)
        print(
            f"   [{idx}] {wh_name or '(无名)'} id={wh_id} "
            f"bundle={bundle_id or '-'} "
            f"state={state or '-'} reason={reason or '-'} "
            f"available={is_avail}"
        )
    if len(warehouses) > 8:
        print(f"   ... 另有 {len(warehouses) - 8} 个仓未列出")


def _rank_available_warehouses(cluster_block):
    warehouses = list(cluster_block.get("warehouses") or [])
    ranked = []
    for wh in warehouses:
        if not wh or not _warehouse_is_available(wh):
            continue
        state = _warehouse_availability_state(wh)
        total_rank = wh.get("total_rank")
        try:
            rank_num = int(total_rank) if total_rank is not None else 9999
        except (TypeError, ValueError):
            rank_num = 9999
        ranked.append((_warehouse_state_rank(state), rank_num, wh, state or "AVAILABLE"))
    ranked.sort(key=lambda x: (x[0], x[1]))
    return ranked


def pick_storage_warehouse_id(cluster_block):
    """从 create/info 的单个集群块中优选 storage/supply warehouse_id（直发）。"""
    ranked = _rank_available_warehouses(cluster_block)
    for _, _, wh, state in ranked:
        wh_id = _warehouse_storage_id(wh)
        if wh_id:
            return wh_id, _warehouse_storage_name(wh), state
    return None, "", ""


def build_cluster_warehouse_entry(cluster_block, mc_id, crossdock=False):
    """
    组装 selected_cluster_warehouses 的单条记录。
    直发用 storage_warehouse_id；中转算仓常返回 storage_warehouse=null，需用 bundle_id。
    """
    ranked = _rank_available_warehouses(cluster_block)
    if not ranked:
        return None, "", ""

    _, _, wh, state = ranked[0]
    storage_id = _warehouse_storage_id(wh)
    bundle_id = (wh.get("bundle_id") or "").strip()
    wh_name = _warehouse_storage_name(wh)
    entry = {"macrolocal_cluster_id": int(mc_id)}

    if crossdock:
        if bundle_id:
            entry["bundle_id"] = bundle_id
            label = wh_name or f"bundle:{bundle_id[:8]}..."
            return entry, label, state
        if storage_id:
            entry["storage_warehouse_id"] = int(storage_id)
            return entry, wh_name, state
        return None, wh_name, state

    if storage_id:
        entry["storage_warehouse_id"] = int(storage_id)
        return entry, wh_name, state
    if bundle_id and (crossdock or wh.get("storage_warehouse") is None):
        entry["bundle_id"] = bundle_id
        label = wh_name or f"bundle:{bundle_id[:8]}..."
        return entry, label, state
    return None, wh_name, state


def resolve_cluster_macrolocal_id(cluster_block, name_by_id, id_to_name):
    """从 create/info 集群块解析 macrolocal_cluster_id。"""
    mc_id = cluster_block.get("macrolocal_cluster_id")
    if mc_id is not None:
        return int(mc_id)
    cluster_id = cluster_block.get("cluster_id")
    if cluster_id is not None:
        cluster_id = int(cluster_id)
        if cluster_id in id_to_name:
            return cluster_id
    cluster_name = (cluster_block.get("cluster_name") or "").strip()
    if cluster_name:
        for name, mid in name_by_id.items():
            if name == cluster_name:
                return int(mid)
    return None


def build_clusters_info_payload(
    cluster_final_items, cluster_id_map, cluster_order=None
):
    """
    组装 multi-cluster/create 的 clusters_info。
    每项为 {macrolocal_cluster_id, items}（items 与 macrolocal_cluster_id 同级）。
    """
    order = cluster_order or sorted(cluster_final_items.keys())
    clusters_info = []
    for cluster_name in order:
        items = cluster_final_items.get(cluster_name)
        if not items:
            continue
        mc_id = cluster_id_map.get(cluster_name)
        if mc_id is None:
            raise ValueError(f"集群 {cluster_name!r} 未解析到 macrolocal_cluster_id")
        clusters_info.append(
            {
                "macrolocal_cluster_id": int(mc_id),
                "items": items,
            }
        )
    if not clusters_info:
        raise ValueError("clusters_info 为空，无法创建多集群草稿")
    return clusters_info


def build_cluster_infos_payload(cluster_final_items, cluster_id_map):
    """兼容旧名，返回 clusters_info。"""
    return build_clusters_info_payload(cluster_final_items, cluster_id_map)


def _skus_by_cluster_id_from_clusters_info(clusters_info):
    """macrolocal_cluster_id -> [sku, ...]"""
    out = {}
    for block in clusters_info or []:
        mc_id = block.get("macrolocal_cluster_id")
        if mc_id is None:
            continue
        skus = []
        for item in block.get("items") or []:
            sku = item.get("sku")
            if sku is not None and sku != "":
                skus.append(sku)
        out[int(mc_id)] = skus
    return out


def poll_multi_cluster_warehouses(
    draft_id, cluster_id_map, headers, proxies, crossdock=False, clusters_info=None
):
    """
    轮询 v2/draft/create/info，为每个集群选仓。
    返回 (selected_list_or_None, error_reason, retry_dropoff)。
    retry_dropoff=False 表示 SKU/矩阵类问题，换交接仓无意义。
    """
    name_by_id = {v: k for k, v in cluster_id_map.items()}
    id_to_name = dict(name_by_id)
    expected_ids = set(cluster_id_map.values())
    sku_by_mc = _skus_by_cluster_id_from_clusters_info(clusters_info)

    while True:
        info_res = ozon_post(
            "https://api-seller.ozon.ru/v2/draft/create/info",
            headers,
            {"draft_id": int(draft_id)},
            proxies,
        )
        current_status = info_res.get("status")
        print(f"📋 多集群算仓状态: [{current_status}]")

        if current_status == "SUCCESS":
            selected = []
            found_ids = set()
            blocks_by_id = {}

            for cluster_block in extract_clusters_from_draft_info(info_res):
                mc_id = resolve_cluster_macrolocal_id(
                    cluster_block, name_by_id, id_to_name
                )
                if mc_id is None:
                    continue
                blocks_by_id[mc_id] = cluster_block

            for mc_id in expected_ids:
                cluster_block = blocks_by_id.get(mc_id)
                cluster_label = name_by_id.get(mc_id, str(mc_id))
                skus = sku_by_mc.get(int(mc_id)) or []
                if not cluster_block:
                    err = f"算仓结果缺少集群: {cluster_label}"
                    print(f"❌ {err}")
                    errors = info_res.get("errors") or []
                    if errors:
                        print(f"   errors: {errors}")
                        err = f"{err}; {_api_err_snippet(errors)}"
                    return None, err, False

                entry, wh_label, state = build_cluster_warehouse_entry(
                    cluster_block, mc_id, crossdock=crossdock
                )
                if not entry:
                    log_cluster_warehouse_diagnostics(cluster_block, cluster_label)
                    err = summarize_cluster_storage_rejection(
                        cluster_block, cluster_label, skus=skus
                    )
                    print(f"❌ {err}")
                    errors = info_res.get("errors") or []
                    if errors:
                        print(f"   errors: {errors}")
                    # 存储仓不可用（尤其 MATRIX）与交接仓无关，勿换 ХАБ 重试
                    return None, err, False

                if entry.get("bundle_id"):
                    print(
                        f"🌟 集群 {cluster_label}: bundle_id={entry['bundle_id']} ({state})"
                    )
                else:
                    print(
                        f"🌟 集群 {cluster_label}: {wh_label} "
                        f"(ID={entry.get('storage_warehouse_id')}, {state})"
                    )
                selected.append(entry)
                found_ids.add(mc_id)

            missing = expected_ids - found_ids
            if missing:
                missing_names = [name_by_id.get(i, i) for i in missing]
                err = f"算仓结果缺少集群: {missing_names}"
                print(f"❌ {err}")
                return None, err, False
            return selected, "", False

        if current_status == "FAILED":
            errors = info_res.get("errors") or info_res
            err = f"多集群算仓失败: {_api_err_snippet(errors)}"
            print(f"❌ {err}")
            err_text = json.dumps(errors, ensure_ascii=False) if not isinstance(errors, str) else errors
            retry = "NOT_AVAILABLE_MATRIX" not in (err_text or "").upper()
            return None, err, retry
        if current_status in ("IN_PROGRESS", "UNSPECIFIED"):
            time.sleep(3)
            continue
        err = f"未知算仓状态: {current_status}"
        print(f"❌ {err}")
        return None, err, False


def create_multi_cluster_draft(
    clusters_info,
    headers,
    proxies,
    delivery_info=None,
):
    """POST /v1/draft/multi-cluster/create，返回 draft_id。"""
    payload = {
        "clusters_info": clusters_info,
        "deletion_sku_mode": "PARTIAL",
    }
    if delivery_info:
        payload["delivery_info"] = delivery_info

    print(
        f"📤 multi-cluster/create: {len(clusters_info)} 个集群, "
        f"delivery_info={'有' if delivery_info else '无'}"
    )
    for idx, block in enumerate(clusters_info):
        item_count = len(block.get("items") or [])
        print(
            f"   集群[{idx}] mc_id={block.get('macrolocal_cluster_id')} "
            f"items={item_count}"
        )
    try:
        draft_res = ozon_post(
            "https://api-seller.ozon.ru/v1/draft/multi-cluster/create",
            headers,
            payload,
            proxies,
        )
    except requests.exceptions.HTTPError as exc:
        resp = getattr(exc, "response", None)
        if resp is not None and is_multi_cluster_client_payload_error(resp.text):
            print("❌ 多集群草稿请求体/items 校验失败，请检查 clusters_info 结构")
        return None
    draft_id = draft_res.get("draft_id")
    if not draft_id:
        print(f"❌ 多集群草稿创建失败: {draft_res}")
        return None
    print(f"✅ 多集群草稿 draft_id={draft_id}")
    return int(draft_id)


def create_direct_draft_and_poll_warehouse(
    cluster_final_items,
    cluster_id_map,
    headers,
    proxies,
):
    """
    单集群直发：v1/draft/direct/create + create/info 选仓。
    直发仅莫斯科单集群，不走 multi-cluster/create（该接口要求 delivery_info）。
    """
    if len(cluster_final_items) != 1:
        names = list(cluster_final_items.keys())
        print(
            f"❌ 直发仅支持单集群（莫斯科），当前有 {len(names)} 个: {names}"
        )
        return None, None, "", (
            f"直发仅支持单集群（莫斯科），当前有 {len(names)} 个: {names}"
        )

    cluster_name = next(iter(cluster_final_items))
    final_items = cluster_final_items[cluster_name]
    mc_id = cluster_id_map.get(cluster_name)
    if mc_id is None:
        err = f"集群 {cluster_name!r} 未解析到 macrolocal_cluster_id"
        print(f"❌ {err}")
        return None, None, "", err

    print("\n📦 [STEP 2] 创建【直发 DIRECT】草稿 (v1/draft/direct/create)...")
    draft_res = ozon_post(
        "https://api-seller.ozon.ru/v1/draft/direct/create",
        headers,
        {
            "cluster_info": {
                "macrolocal_cluster_id": int(mc_id),
                "items": final_items,
            },
            "deletion_sku_mode": "PARTIAL",
        },
        proxies,
    )
    draft_id = draft_res.get("draft_id")
    if not draft_id:
        err = f"直发草稿创建失败: {_api_err_snippet(draft_res)}"
        print(f"❌ {err}")
        return None, None, "", err
    print(f"✅ 直发草稿 draft_id={draft_id}")

    print("\n⏳ [STEP 3] 直发算仓...")
    while True:
        info_res = ozon_post(
            "https://api-seller.ozon.ru/v2/draft/create/info",
            headers,
            {"draft_id": int(draft_id)},
            proxies,
        )
        current_status = info_res.get("status")
        print(f"📋 直发算仓状态: [{current_status}]")

        if current_status == "SUCCESS":
            blocks = list(extract_clusters_from_draft_info(info_res))
            for cluster_block in blocks:
                entry, wh_label, state = build_cluster_warehouse_entry(
                    cluster_block, mc_id, crossdock=False
                )
                if entry:
                    wh_id = entry.get("storage_warehouse_id")
                    print(
                        f"🌟 集群 {cluster_name}: {wh_label} "
                        f"(ID={wh_id}, {state})"
                    )
                    return int(draft_id), [entry], wh_label, ""
            if blocks:
                log_cluster_warehouse_diagnostics(blocks[0], cluster_name)
            err = f"集群 {cluster_name} 无可用 FBO 仓库（算仓 SUCCESS 但无可用仓）"
            print(f"❌ {err}")
            return None, None, "", err
        if current_status == "FAILED":
            err = f"直发算仓失败: {_api_err_snippet(info_res.get('errors') or info_res)}"
            print(f"❌ {err}")
            return None, None, "", err
        if current_status in ("IN_PROGRESS", "UNSPECIFIED"):
            time.sleep(3)
            continue
        err = f"直发算仓未知状态 {current_status!r}: {_api_err_snippet(info_res)}"
        print(f"❌ {err}")
        return None, None, "", err


def build_cluster_crossdock_delivery_attempts(
    cluster_id_map,
    headers,
    proxies,
    default_dropoff_name=DEFAULT_CROSSDOCK_DROP_OFF_NAME,
    nearby_max_km=80,
):
    """各集群首选交接仓 → multi-cluster/create 的 delivery_info 尝试序列。"""
    attempts = []
    seen_hub_ids = set()

    def add_attempt(cluster_name, hub):
        wid = int(hub["warehouse_id"])
        if wid in seen_hub_ids:
            return
        seen_hub_ids.add(wid)
        attempts.append((cluster_name, hub))

    cluster_order = sorted(
        cluster_id_map.keys(),
        key=lambda name: (0 if is_moscow_cluster(name) else 1, name),
    )
    for cluster_name in cluster_order:
        mc_id = cluster_id_map[cluster_name]
        merged = merge_dropoff_candidates(
            fetch_cluster_dropoff_candidates(headers, proxies, mc_id),
            fetch_fbo_crossdock_dropoffs(
                headers,
                proxies,
                fbo_search_keywords_for_name(default_dropoff_name),
            ),
        )
        try_order = build_dropoff_try_order(
            merged, default_dropoff_name, nearby_max_km
        )
        if try_order:
            print(
                f"   集群 {cluster_name} 首选交接仓: {try_order[0]['name']} "
                f"(ID={try_order[0]['warehouse_id']})"
            )
            add_attempt(cluster_name, try_order[0])
            for alt in try_order[1:16]:
                add_attempt(cluster_name, alt)

    if not attempts:
        merged = merge_dropoff_candidates(
            fetch_all_clusters_dropoff_candidates(
                headers, proxies, cluster_id_map
            ),
            fetch_fbo_crossdock_dropoffs(
                headers,
                proxies,
                fbo_search_keywords_for_name(default_dropoff_name),
            ),
        )
        for hub in build_dropoff_try_order(
            merged, default_dropoff_name, nearby_max_km
        )[:5]:
            add_attempt("合并", hub)
    return attempts


def create_multi_cluster_crossdock_draft(
    clusters_info,
    cluster_id_map,
    headers,
    proxies,
    drop_off_warehouse_name=DEFAULT_CROSSDOCK_DROP_OFF_NAME,
    drop_off_nearby_max_km=80,
    drop_off_retry_delay=6,
    crossdock_delivery_type="DROPOFF",
    seller_warehouse_id=0,
):
    """
    中转多集群：优先全部集群共用同一个 delivery_info（统一发运）。
    Ozon multi-cluster/create 仅支持顶层一个交接仓；算仓失败时再换附近 ХАБ。
    转正时各集群用 create/info 返回的 bundle_id（storage_warehouse 常为 null）。
    """
    delivery_attempts = build_cluster_crossdock_delivery_attempts(
        cluster_id_map,
        headers,
        proxies,
        default_dropoff_name=drop_off_warehouse_name,
        nearby_max_km=drop_off_nearby_max_km,
    )

    print(
        f"🚚 多集群中转：{len(clusters_info)} 个集群，"
        f"将依次尝试 {len(delivery_attempts)} 个交接仓方案（草稿层仅 1 个 delivery_info）"
    )

    last_detail = ""
    for idx, (cluster_label, wh) in enumerate(delivery_attempts, start=1):
        dist_hint = (
            f" 距离首选 {wh['distance_km']}km"
            if wh.get("distance_km") is not None
            else ""
        )
        print(
            f"\n🔄 多集群草稿尝试 [{idx}/{len(delivery_attempts)}]: "
            f"参考集群={cluster_label}  交接仓={wh['name']} "
            f"(ID={wh['warehouse_id']}){dist_hint}"
        )
        if idx > 1:
            print(f"⏳ 频率保护挂起 {drop_off_retry_delay}s...")
            time.sleep(drop_off_retry_delay)

        delivery_info = (
            {"type": "PICKUP", "seller_warehouse_id": int(seller_warehouse_id)}
            if crossdock_delivery_type.upper() == "PICKUP"
            else {
                "type": "DROPOFF",
                "drop_off_warehouse": {
                    "warehouse_id": int(wh["warehouse_id"]),
                    "warehouse_type": wh["warehouse_type"],
                },
            }
        )
        draft_id = create_multi_cluster_draft(
            clusters_info, headers, proxies, delivery_info=delivery_info
        )
        if not draft_id:
            last_detail = f"交接仓 {wh['name']} 创建草稿失败"
            continue

        selected, poll_err, retry_dropoff = poll_multi_cluster_warehouses(
            draft_id,
            cluster_id_map,
            headers,
            proxies,
            crossdock=True,
            clusters_info=clusters_info,
        )
        if selected:
            print(
                f"✅ 多集群草稿算仓通过（草稿 delivery_info 交接仓: {wh['name']}）"
            )
            return draft_id, selected, wh["name"], ""

        last_detail = poll_err or f"交接仓 {wh['name']} 算仓未通过"
        if not retry_dropoff:
            # SKU 不在集群矩阵等：换交接仓无效，直接返回真实原因
            print(f"❌ {last_detail}（换交接仓无效，停止重试）")
            return None, None, "", last_detail
        print(f"⚠️ 交接仓 {wh['name']} 算仓未通过，尝试下一方案...")
    if last_detail:
        err = (
            f"多集群中转草稿失败：已尝试 {len(delivery_attempts)} 个交接仓方案。"
            f"末次原因: {last_detail}"
        )
    else:
        err = (
            f"多集群中转草稿失败：已尝试 {len(delivery_attempts)} 个交接仓方案，"
            "算仓均未通过"
        )
    return None, None, "", err


def build_global_ordered_cargo_ids(shipment_items, cargo_ids_by_cluster):
    """
    按 shipment_items（唯一ID 全局顺序）拼接各集群 supply 的 cargo_id 列表。
    每行 ITEM 按 quantity//number 展开为多箱，与 build_cargoes_from_items 一致。
    cargo_ids_by_cluster: {cluster_name: [cargo_id, ...]}
    """
    shipment_items = sort_shipment_items_by_row_id(shipment_items)
    normalized_by_cluster = {
        cluster: [int(cid) for cid in ids]
        for cluster, ids in cargo_ids_by_cluster.items()
    }
    indices = defaultdict(int)
    ordered = []
    for row in shipment_items:
        cluster = row.get("cluster") or "未知集群"
        per_box = int(row["number"])
        qty = int(row["quantity"])
        if per_box <= 0 or qty % per_box != 0:
            raise ValueError(
                f"货号 {row.get('offer_id') or row.get('sku')} "
                f"数量 {qty} 无法按单箱 {per_box} 整除"
            )
        cluster_ids = normalized_by_cluster.get(cluster) or []
        box_count = qty // per_box
        for _ in range(box_count):
            idx = indices[cluster]
            if idx >= len(cluster_ids):
                raise ValueError(
                    f"集群 {cluster!r} cargo_id 不足: 需要第 {idx + 1} 个，"
                    f"仅有 {len(cluster_ids)} 个"
                )
            ordered.append(cluster_ids[idx])
            indices[cluster] += 1
    return ordered


def build_global_ordered_cargo_ids_tagged(shipment_items, cargo_ids_by_method_cluster):
    """
    跨直发/中转合并时按 (发货方式, 集群) 取 cargo_id。
    每行 ITEM 按 quantity//number 展开为多箱。
    cargo_ids_by_method_cluster: {(method, cluster): [cargo_id, ...]}
    """
    shipment_items = sort_shipment_items_by_row_id(shipment_items)
    normalized = {
        key: [int(cid) for cid in ids]
        for key, ids in cargo_ids_by_method_cluster.items()
    }
    indices = defaultdict(int)
    ordered = []
    for row in shipment_items:
        method = row.get("_ship_method") or row.get("shipping_method") or ""
        cluster = row.get("cluster") or "未知集群"
        per_box = int(row["number"])
        qty = int(row["quantity"])
        if per_box <= 0 or qty % per_box != 0:
            raise ValueError(
                f"货号 {row.get('offer_id') or row.get('sku')} "
                f"数量 {qty} 无法按单箱 {per_box} 整除"
            )
        key = (method, cluster)
        cluster_ids = normalized.get(key) or []
        box_count = qty // per_box
        for _ in range(box_count):
            idx = indices[key]
            if idx >= len(cluster_ids):
                raise ValueError(
                    f"{method}/{cluster!r} cargo_id 不足: 需要第 {idx + 1} 个，"
                    f"仅有 {len(cluster_ids)} 个"
                )
            ordered.append(cluster_ids[idx])
            indices[key] += 1
    return ordered


def tag_shipment_items_with_method(shipment_items, shipping_method):
    tagged = []
    for row in shipment_items:
        item = dict(row)
        item["_ship_method"] = shipping_method
        tagged.append(item)
    return tagged


def cluster_names_in_item_order(shipment_items):
    """按唯一ID顺序返回集群名首次出现序列（去重）。"""
    seen = []
    for row in sort_shipment_items_by_row_id(shipment_items):
        name = row.get("cluster") or "未知集群"
        if name not in seen:
            seen.append(name)
    return seen


def extract_pdf_bytes_from_label_file(file_bytes, content_type=None):
    """从 Ozon 箱唛下载结果提取 PDF 字节（支持 zip 多文件合并为单 PDF）。"""
    if not file_bytes:
        return None
    if file_bytes[:4] == b"%PDF":
        return file_bytes

    is_zip = (
        zipfile.is_zipfile(io.BytesIO(file_bytes))
        or "zip" in (content_type or "").lower()
    )
    if not is_zip:
        return file_bytes

    pdf_parts = []
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        for name in sorted(zf.namelist()):
            if name.lower().endswith(".pdf"):
                pdf_parts.append(zf.read(name))
    if not pdf_parts:
        raise ValueError("箱唛 zip 中未找到 PDF 文件")
    if len(pdf_parts) == 1:
        return pdf_parts[0]

    try:
        import fitz
    except ImportError as e:
        raise ImportError(
            "需要安装 PyMuPDF (pip install pymupdf) 才能合并 zip 内多 PDF"
        ) from e

    merged = fitz.open()
    for part in pdf_parts:
        src = fitz.open(stream=part, filetype="pdf")
        for page_idx in range(src.page_count):
            _append_normalized_label_page(merged, src, page_idx)
        src.close()
    out = merged.tobytes()
    merged.close()
    return out


def prepare_supply_label_pdf_bytes(
    supply_id, cluster_cargo_ids, headers, proxies
):
    """
    下载单个 supply 箱唛，解压 zip（如有），并按该 supply 内货位顺序重排页序。
    """
    raw_bytes, content_type = fetch_box_label_pdf_bytes(
        supply_id, headers, proxies, ordered_cargo_ids=cluster_cargo_ids
    )
    if not raw_bytes:
        return None
    pdf_bytes = extract_pdf_bytes_from_label_file(raw_bytes, content_type)
    if not pdf_bytes or pdf_bytes[:4] != b"%PDF":
        raise ValueError(f"supply_id={supply_id} 箱唛不是有效 PDF")
    cargo_ids = [int(cid) for cid in cluster_cargo_ids]
    return reorder_cargoes_label_pdf(pdf_bytes, cargo_ids)


def export_merged_labels_and_excel(
    order_dir,
    shipment_items,
    cargo_ids_by_cluster,
    supply_by_cluster,
    headers,
    proxies,
    internal_order_no,
    write_excel=True,
    unified_supply=False,
    batch_or_order=None,
):
    """
    多集群：各 supply 箱唛 → 按唯一ID全局顺序合并为 1 本 PDF；
    顺序表/询价表与合并 PDF 使用同一 cargo_id 顺序。
    write_excel=False 时仅下载箱唛，供发货人+店铺级总表合并。
    """
    shipment_items = sort_shipment_items_by_row_id(shipment_items)
    global_ordered_cargo_ids = build_global_ordered_cargo_ids(
        shipment_items, cargo_ids_by_cluster
    )
    print(
        f"\n🏷️ 下载箱唛并按唯一ID顺序合并"
        f"（共 {len(global_ordered_cargo_ids)} 箱）..."
    )

    pdf_bytes_list = []
    if unified_supply:
        supply_id = next(iter(supply_by_cluster.values()))["supply_id"]
        print(
            f"   单 supply_id={supply_id} "
            f"({len(global_ordered_cargo_ids)} 箱)..."
        )
        ordered_pdf = prepare_supply_label_pdf_bytes(
            supply_id, global_ordered_cargo_ids, headers, proxies
        )
        if not ordered_pdf:
            raise ValueError(f"supply_id={supply_id} 箱唛下载失败")
        raw_path = os.path.join(order_dir, "箱唛_合并supply.pdf")
        with open(raw_path, "wb") as f:
            f.write(ordered_pdf)
        print(f"   💾 合并 supply 箱唛: {raw_path}")
        pdf_bytes_list.append(ordered_pdf)
        merged_bytes = ordered_pdf
    else:
        for cluster_name in cluster_names_in_item_order(shipment_items):
            sup = supply_by_cluster.get(cluster_name)
            if not sup:
                raise ValueError(f"集群 {cluster_name!r} 无 supply 映射")
            supply_id = sup["supply_id"]
            cluster_cargo_ids = cargo_ids_by_cluster[cluster_name]
            print(
                f"   集群 {cluster_name} supply_id={supply_id} "
                f"({len(cluster_cargo_ids)} 箱)..."
            )
            ordered_pdf = prepare_supply_label_pdf_bytes(
                supply_id, cluster_cargo_ids, headers, proxies
            )
            if not ordered_pdf:
                raise ValueError(f"集群 {cluster_name} 箱唛下载失败")
            raw_path = os.path.join(
                order_dir, f"箱唛_{sanitize_folder_name(cluster_name)}.pdf"
            )
            with open(raw_path, "wb") as f:
                f.write(ordered_pdf)
            print(f"   💾 集群箱唛: {raw_path}")
            pdf_bytes_list.append(ordered_pdf)

        merged_bytes = merge_multi_supply_label_pdfs(
            pdf_bytes_list, global_ordered_cargo_ids
        )
    ordered_path = os.path.join(order_dir, "交货货位标签_按ITEMS顺序.pdf")
    with open(ordered_path, "wb") as f:
        f.write(merged_bytes)
    print(f"✅ 已按唯一ID顺序合并箱唛 PDF: {ordered_path}")

    if write_excel:
        box_meta = build_box_meta_from_items(shipment_items)
        full_box_meta = enrich_box_meta_with_cargo_ids(
            box_meta, global_ordered_cargo_ids
        )
        export_batch_excel_reports(
            order_dir,
            full_box_meta,
            internal_order_no,
            batch_or_order=batch_or_order,
        )
    return ordered_path, global_ordered_cargo_ids, pdf_bytes_list


def finalize_combined_exports(
    order_dir, export_bundles, internal_order_no, batch_or_order=None
):
    """
    同一发货人+店铺下，合并直发/中转的总箱唛、询价表、顺序表（按唯一ID全局顺序）。
    """
    if not export_bundles:
        return None, None, None

    os.makedirs(order_dir, exist_ok=True)
    all_items = []
    cargo_ids_by_method_cluster = {}
    pdf_bytes_list = []

    for bundle in export_bundles:
        method = bundle["shipping_method"]
        for row in bundle.get("shipment_items") or []:
            item = dict(row)
            item["_ship_method"] = method
            all_items.append(item)
        for cluster, ids in (bundle.get("cargo_ids_by_cluster") or {}).items():
            cargo_ids_by_method_cluster[(method, cluster)] = ids
        pdf_bytes_list.extend(bundle.get("cluster_pdf_bytes") or [])

    all_items = sort_shipment_items_by_row_id(all_items)
    global_ordered_cargo_ids = build_global_ordered_cargo_ids_tagged(
        all_items, cargo_ids_by_method_cluster
    )

    print(
        f"\n📦 生成发货人+店铺总输出（{len(export_bundles)} 个 order，"
        f"共 {len(global_ordered_cargo_ids)} 箱）→ {order_dir}"
    )
    merged_bytes = merge_multi_supply_label_pdfs(
        pdf_bytes_list, global_ordered_cargo_ids
    )
    ordered_path = os.path.join(order_dir, "交货货位标签_按ITEMS顺序.pdf")
    with open(ordered_path, "wb") as f:
        f.write(merged_bytes)
    print(f"✅ 总箱唛 PDF: {ordered_path}")

    box_meta = build_box_meta_from_items(all_items)
    full_box_meta = enrich_box_meta_with_cargo_ids(
        box_meta, global_ordered_cargo_ids
    )
    export_batch_excel_reports(
        order_dir,
        full_box_meta,
        internal_order_no,
        batch_or_order=batch_or_order,
    )
    return ordered_path, global_ordered_cargo_ids, full_box_meta


def fetch_box_label_pdf_bytes(
    supply_id, headers, proxies, ordered_cargo_ids=None
):
    """生成并下载箱唛 PDF，返回 (file_bytes, content_type)。"""
    cargo_ids = list(ordered_cargo_ids or [])
    label_create_payload = {"supply_id": int(supply_id)}
    if cargo_ids:
        label_create_payload["cargoes"] = [{"cargo_id": cid} for cid in cargo_ids]

    label_create_res = ozon_post(
        "https://api-seller.ozon.ru/v1/cargoes-label/create",
        headers,
        label_create_payload,
        proxies,
    )
    if label_create_res.get("errors"):
        print(f"❌ 箱唛生成任务创建失败: {label_create_res}")
        return None, None

    label_operation_id = label_create_res.get("operation_id")
    if not label_operation_id:
        print(f"❌ 未获得箱唛任务 operation_id: {label_create_res}")
        return None, None

    file_url = file_guid = file_content_b64 = None
    while True:
        label_get_res = ozon_post(
            "https://api-seller.ozon.ru/v1/cargoes-label/get",
            headers,
            {"operation_id": label_operation_id},
            proxies,
        )
        label_status = label_get_res.get("status", "")
        if label_status == "SUCCESS":
            file_url, file_guid, file_content_b64 = extract_label_file_info(
                label_get_res
            )
            break
        if label_status == "FAILED":
            print(f"❌ 箱唛生成失败: {label_get_res}")
            return None, None
        time.sleep(2)

    try:
        file_bytes, content_type, _ = download_box_label_content(
            file_url, file_guid, file_content_b64, headers, proxies
        )
    except (requests.exceptions.HTTPError, ValueError) as e:
        print(f"❌ 箱唛下载失败: {e}")
        return None, None
    return file_bytes, content_type


def merge_multi_supply_label_pdfs(pdf_bytes_list, global_ordered_cargo_ids):
    """将多个 supply 的箱唛 PDF 按全局 cargo_id 顺序合并为一本 PDF。"""
    try:
        import fitz
    except ImportError as e:
        raise ImportError(
            "需要安装 PyMuPDF (pip install pymupdf) 才能合并箱唛 PDF"
        ) from e

    page_by_cargo = {}
    expected_set = set(global_ordered_cargo_ids)
    src_docs = []

    for pdf_bytes in pdf_bytes_list:
        if not pdf_bytes:
            continue
        src = fitz.open(stream=pdf_bytes, filetype="pdf")
        src_docs.append(src)
        for page_idx in range(src.page_count):
            cid = _page_cargo_id(src[page_idx], expected_set)
            if cid is not None and cid not in page_by_cargo:
                page_by_cargo[cid] = (len(src_docs) - 1, page_idx)

    missing = [cid for cid in global_ordered_cargo_ids if cid not in page_by_cargo]
    if missing:
        for doc in src_docs:
            doc.close()
        raise ValueError(
            f"合并箱唛 PDF 时缺少 {len(missing)} 个货位页面: {missing[:3]}..."
        )

    dst = fitz.open()
    for cid in global_ordered_cargo_ids:
        doc_idx, page_idx = page_by_cargo[cid]
        _append_normalized_label_page(dst, src_docs[doc_idx], page_idx)

    merged_bytes = dst.tobytes()
    dst.close()
    for doc in src_docs:
        doc.close()
    return merged_bytes


def submit_cargoes_for_supply(
    supply_id,
    cluster_shipment_items,
    headers,
    proxies,
    order_id,
    batch_no,
    shop,
):
    """对单个 supply 提交货位，返回 (ok, ordered_cargo_ids)。"""
    cargoes_list = build_cargoes_from_items(cluster_shipment_items)
    box_count = len(cargoes_list)
    transport_cargo_id = ensure_large_supply_transport_setup(
        supply_id, box_count, headers, proxies
    )
    if box_count > MAX_BOXES_PER_SUPPLY_ORDER and not transport_cargo_id:
        return False, [], "运输货位初始化失败（大货量需 activate transport）"

    cargoes_ok, ordered_cargo_ids, cargo_err = submit_cargoes_in_batches(
        supply_id,
        cargoes_list,
        headers,
        proxies,
        order_id=order_id,
        batch_no=batch_no,
        shop=shop,
        full_total=box_count,
    )
    if not cargoes_ok:
        return False, [], cargo_err or "货位分批提交失败"

    if len(ordered_cargo_ids) != len(cargoes_list):
        ordered_cargo_ids = fetch_ordered_box_cargo_ids(
            supply_id, cargoes_list, headers, proxies
        )

    if transport_cargo_id:
        box_cargo_ids = fetch_ordered_box_cargo_ids(
            supply_id, cargoes_list, headers, proxies
        )
        if not bind_boxes_to_transport_cargo(
            supply_id, transport_cargo_id, box_cargo_ids, headers, proxies
        ):
            return False, [], "盒子货位绑定运输货位失败"

    return True, ordered_cargo_ids, ""


# ==================== 主业务流水线 ====================
def run_single_application(
    target_shop,
    shipping_method,
    macrolocal_cluster,
    items,
    batch_no,
    crossdock_delivery_type="DROPOFF",
    drop_off_warehouse_name=DEFAULT_CROSSDOCK_DROP_OFF_NAME,
    drop_off_nearby_max_km=80,
    drop_off_retry_delay=6,
    seller_warehouse_id=0,
    internal_order_no="",
    shipper="",
    archive_date=None,
):
    """执行一次 Ozon 发货全流程，成功返回 order_id，失败返回 None。"""
    if target_shop not in SHOP_DATA:
        print(f"❌ 错误：在授权表中未找到店铺【{target_shop}】的配置，请核对！")
        return None

    client_id = SHOP_DATA[target_shop]["client_id"]
    api_key = SHOP_DATA[target_shop]["api_key"]
    proxies = {"http": None, "https": None}
    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }

    print("=" * 60)
    print(f"📋 批次号: {batch_no}")
    print(f"🏪 店铺: 【{target_shop}】 (Client-Id: {client_id})")
    print(f"🌍 集群: {macrolocal_cluster}")
    print(f"🚀 发货方式: 【{shipping_method}】")
    print(f"📦 ITEMS: {items}")
    print("=" * 60)

    print(f"🔍 [STEP 0] 正在检索宏观集群主键 ID: {macrolocal_cluster!r}")
    try:
        macrolocal_cluster_id = resolve_macrolocal_cluster_id(
            macrolocal_cluster, headers, proxies
        )
        print(f"✅ 成功锁定大区集群 ID = {macrolocal_cluster_id}")
    except ValueError as e:
        print(f"❌ 运行中止: {e}")
        return None

    print("\n🔍 [STEP 1] 开始批量查询货号(offer_id)对应的真实数字 SKU...")
    offer_ids = list(dict.fromkeys(item["sku"] for item in items))
    sku_data = ozon_post(
        "https://api-seller.ozon.ru/v3/product/info/list",
        headers,
        {"offer_id": offer_ids},
        proxies,
    )

    offer_to_sku_map = {}
    offer_to_barcode_map = {}
    for p in sku_data.get("items", []):
        real_sku = p.get("sku")
        oid = p.get("offer_id")
        offer_to_sku_map[oid] = real_sku
        barcodes = p.get("barcodes") or []
        if barcodes:
            offer_to_barcode_map[oid] = barcodes[0]

    shipment_items = []
    draft_qty_by_sku = {}
    for item in items:
        user_offer_id = item["sku"]
        real_numeric_sku = offer_to_sku_map.get(user_offer_id)
        if not real_numeric_sku:
            print(
                f"⚠️ 警告: 货号 '{user_offer_id}' 在后台未查到任何有效的数字 SKU！该商品将被跳过。"
            )
            continue
        sku_int = int(real_numeric_sku)
        draft_qty_by_sku[sku_int] = (
            draft_qty_by_sku.get(sku_int, 0) + item["quantity"]
        )
        shipment_items.append(
            {
                "offer_id": user_offer_id,
                "quantity": item["quantity"],
                "number": item["number"],
                "barcode": offer_to_barcode_map.get(user_offer_id),
                **item_meta_kwargs(item),
            }
        )

    final_items = [
        {"sku": sku, "quantity": qty} for sku, qty in draft_qty_by_sku.items()
    ]
    if not final_items:
        print("❌ 错误: 没有任何可以成功转换的 SKU，流程强行终止。")
        return None

    total_boxes = sum(
        row["quantity"] // row["number"]
        for row in shipment_items
        if row["number"] > 0 and row["quantity"] % row["number"] == 0
    )
    print(
        f"✅ 核心数据解析就绪。预计拆装箱数: {total_boxes} 箱。商品总清单: {final_items}"
    )
    if total_boxes > MAX_BOXES_PER_SUPPLY_ORDER:
        print(
            f"📦 超过 {MAX_BOXES_PER_SUPPLY_ORDER} 箱："
            f"activate → 创建 1 个运输货位 → 分批提交盒子 → bind 绑定"
        )
    time.sleep(1.5)

    # ===================================================
    # [STEP 2+3] 依据业务逻辑执行核心草稿构建与仓位闭环
    # ===================================================
    draft_id = None
    target_warehouse_id = None
    cluster_warehouse_extra = {}
    final_dropoff_wh_name = ""

    if shipping_method == "直发":
        print("\n📦 [STEP 2] 正在创建【官方直发(DIRECT)】草稿单...")
        draft_res = ozon_post(
            "https://api-seller.ozon.ru/v1/draft/direct/create",
            headers,
            {
                "cluster_info": {
                    "macrolocal_cluster_id": macrolocal_cluster_id,
                    "items": final_items,
                },
                "deletion_sku_mode": "PARTIAL",
            },
            proxies,
        )
        draft_id = draft_res.get("draft_id")
        if not draft_id:
            print("❌ 错误: 官方直发草稿创建失败！")
            return None

        print(f"\n⏳ [STEP 3] 正在轮询计算直发接收仓库...")
        while True:
            info_res = ozon_post(
                "https://api-seller.ozon.ru/v2/draft/create/info",
                headers,
                {"draft_id": draft_id},
                proxies,
            )
            current_status = info_res.get("status")
            print(f"📋 直发算仓状态: [{current_status}]")

            if current_status == "SUCCESS":
                target_warehouse_id = None
                for c in extract_clusters_from_draft_info(info_res):
                    wh_id, wh_name, state = pick_storage_warehouse_id(c)
                    if wh_id:
                        target_warehouse_id = wh_id
                        if "PARTIAL" in (state or "").upper():
                            print(
                                f"⚠️ 降级选仓: {wh_name} (ID: {target_warehouse_id}, {state})"
                            )
                        else:
                            print(
                                f"🌟 选仓成功: {wh_name} (ID: {target_warehouse_id}, {state})"
                            )
                        break
                break
            elif current_status == "FAILED":
                print("❌ 官方直发算法判定库存或类目无法落仓。")
                return None
            time.sleep(3)

    elif shipping_method == "中转":
        print("\n📦 [STEP 2+3] 启动【越库中转(CROSSDOCK)】多仓并联重试引擎...")
        cluster_cands = fetch_cluster_dropoff_candidates(
            headers, proxies, macrolocal_cluster_id
        )
        fbo_cands = fetch_fbo_crossdock_dropoffs(
            headers, proxies, fbo_search_keywords_for_name(drop_off_warehouse_name)
        )
        dropoff_candidates = merge_dropoff_candidates(cluster_cands, fbo_cands)
        dropoff_try_order = build_dropoff_try_order(
            dropoff_candidates,
            drop_off_warehouse_name,
            drop_off_nearby_max_km,
        )

        for idx, wh in enumerate(dropoff_try_order, start=1):
            dist_hint = (
                f" 距离首选 {wh['distance_km']}km"
                if wh.get("distance_km") is not None
                else ""
            )
            print(
                f"\n🔄 尝试交接仓方案 [{idx}/{len(dropoff_try_order)}]: {wh['name']} (ID={wh['warehouse_id']}){dist_hint}"
            )
            if idx > 1:
                print(
                    f"⏳ 触发频率保护安全锁，强制挂起 {drop_off_retry_delay}s..."
                )
                time.sleep(drop_off_retry_delay)

            draft_id, warehouse_extra, _, info_status = (
                create_crossdock_draft_and_poll_storage(
                    wh,
                    final_items,
                    macrolocal_cluster_id,
                    crossdock_delivery_type,
                    seller_warehouse_id,
                    headers,
                    proxies,
                )
            )
            if (
                draft_id
                and info_status == "SUCCESS"
                and warehouse_extra
            ):
                cluster_warehouse_extra = warehouse_extra
                final_dropoff_wh_name = wh["name"]
                print(f"✅ 成功锁定中转链路目标交接仓: {wh['name']}")
                break
        else:
            print(
                "❌ 错误: 备选的所有中转交接仓均由于时段或库容拦截算仓失败。"
            )
            return None

    if not draft_id:
        print("❌ 核心业务草稿未正常建立，程序阻断。")
        return None
    time.sleep(2)

    # ===================================================
    # [STEP 4] 供货方案申请转正确认
    # ===================================================
    print(f"\n🚀 [STEP 4] 正在执行方案转正申报 (v2/draft/supply/create)...")
    supply_url = "https://api-seller.ozon.ru/v2/draft/supply/create"
    cluster_entry = {"macrolocal_cluster_id": macrolocal_cluster_id}
    if shipping_method == "直发" and target_warehouse_id:
        cluster_entry["storage_warehouse_id"] = int(target_warehouse_id)
    elif shipping_method == "中转":
        cluster_entry.update(cluster_warehouse_extra)

    supply_payload = {
        "draft_id": int(draft_id),
        "selected_cluster_warehouses": [cluster_entry],
        "supply_type": "DIRECT" if shipping_method == "直发" else "CROSSDOCK",
    }

    if shipping_method == "中转":
        print("📅 未配置 timeslot（遵照指示：不用选择发货时间）")

    supply_res = ozon_post(supply_url, headers, supply_payload, proxies)

    # ===================================================
    # [STEP 5] 异步轮询获取唯一业务订货单 order_id
    # ===================================================
    print("\n⏳ [STEP 5] 正在向官方异步追溯生成的发货订单主键 order_id...")
    status_payload = {"draft_id": draft_id}
    if supply_res.get("operation_id"):
        status_payload["operation_id"] = supply_res["operation_id"]

    real_order_id = None
    while True:
        status_res = ozon_post(
            "https://api-seller.ozon.ru/v2/draft/supply/create/status",
            headers,
            status_payload,
            proxies,
        )
        current_status = status_res.get("status", "")
        candidate_order_id = status_res.get("order_id") or 0

        if candidate_order_id and int(candidate_order_id) > 0:
            real_order_id = int(candidate_order_id)
            break
        order_ids = status_res.get("result", {}).get("order_ids", [])
        if order_ids:
            real_order_id = int(order_ids[0])
            break

        if str(current_status).upper() in (
            "FAILED",
            "DRAFTSUPPLYCREATESTATUSFAILED",
        ):
            print(f"❌ 方案确认失败: {status_res}")
            return None
        time.sleep(3)

    print(f"✅ 成功斩获发货申请单 order_id: {real_order_id}")
    batch_or_order = (str(batch_no).strip() if batch_no else "") or internal_order_no or "未知单号"
    order_dir = build_order_output_dir(
        shipper,
        target_shop,
        batch_or_order,
        archive_date=archive_date,
    )
    os.makedirs(order_dir, exist_ok=True)
    print(f"📁 输出目录: {order_dir}")

    # ===================================================
    # [STEP 6] 解析货件批次 supply_id
    # ===================================================
    print("\n🔎 [STEP 6] 正在抓取 supply_id...")
    order_detail_res = ozon_post(
        "https://api-seller.ozon.ru/v3/supply-order/get",
        headers,
        {"order_ids": [str(real_order_id)]},
        proxies,
    )
    real_supply_id = extract_supply_id_from_order_detail(order_detail_res)
    if not real_supply_id:
        print("❌ 致命错误: 未能通过详单获取到 supply_id")
        return None
    print(f"✅ 成功锁定全局货运批次号 supply_id: {real_supply_id}")
    time.sleep(1.5)

    cargoes_list = build_cargoes_from_items(shipment_items)
    box_meta = build_box_meta_from_items(shipment_items)
    transport_cargo_id = ensure_large_supply_transport_setup(
        real_supply_id,
        len(cargoes_list),
        headers,
        proxies,
    )
    if (
        len(cargoes_list) > MAX_BOXES_PER_SUPPLY_ORDER
        and not transport_cargo_id
    ):
        print("❌ 运输货位初始化失败，终止本批次。")
        return None
    time.sleep(1.5)

    # ===================================================
    # [STEP 7] 提交一箱一 SKU 箱位舱单
    # ===================================================
    print("\n📦 [STEP 7] 正在提交货件信息（грузоместа）...")
    cargoes_ok, ordered_cargo_ids, _ = submit_cargoes_in_batches(
        real_supply_id,
        cargoes_list,
        headers,
        proxies,
        order_id=real_order_id,
        batch_no=batch_no,
        shop=target_shop,
        full_total=len(cargoes_list),
    )
    if not cargoes_ok:
        print("❌ 货位提交失败，终止本批次。")
        return None
    if len(ordered_cargo_ids) != len(cargoes_list):
        ordered_cargo_ids = fetch_ordered_box_cargo_ids(
            real_supply_id, cargoes_list, headers, proxies
        )

    if transport_cargo_id:
        box_cargo_ids = fetch_ordered_box_cargo_ids(
            real_supply_id, cargoes_list, headers, proxies
        )
        if not bind_boxes_to_transport_cargo(
            real_supply_id,
            transport_cargo_id,
            box_cargo_ids,
            headers,
            proxies,
        ):
            print("❌ 运输货位绑定失败，终止本批次。")
            return None
    time.sleep(1.5)

    # ===================================================
    # [STEP 8] 下载箱唛、按 ITEMS 重排、写入对应关系
    # ===================================================
    box_label_path = download_box_labels(
        real_supply_id,
        order_dir,
        headers,
        proxies,
        ordered_cargo_ids=ordered_cargo_ids,
        box_meta=box_meta,
    )
    if not box_label_path:
        print("❌ 箱唛下载失败，终止本批次。")
        return None

    full_box_meta = enrich_box_meta_with_cargo_ids(
        box_meta, ordered_cargo_ids
    )
    try:
        export_batch_excel_reports(
            order_dir,
            full_box_meta,
            internal_order_no,
            batch_or_order=batch_or_order,
        )
    except Exception as exc:
        print(f"⚠️ Excel 导出失败: {exc}")

    inquiry_name, order_name = build_excel_report_filenames(
        batch_or_order, len(full_box_meta)
    )

    # ===================================================
    # 🎉 数据终端看板
    # ===================================================
    print(f"\n==================================================")
    print(f"🎉 🎉 【全流程大一统脚本执行完成】 🎉 🎉")
    print(f"⚙️ 业务流转选择: {shipping_method}模式")
    print(f"📦 草稿 draft_id: {draft_id}")
    print(f"🆔 供货单 order_id: {real_order_id}")
    print(f"🚗 交货批次 supply_id: {real_supply_id}")
    if shipping_method == "中转":
        print(f"🚚 交接仓: {final_dropoff_wh_name}")
    print(f"⚙️ FBO 仓库 ID: {target_warehouse_id or '（越库无需指定）'}")
    print(f"📦 箱位数: {len(cargoes_list)}")
    print(f"📁 输出目录: {order_dir}")
    print(f"   ├─ 原始箱唛 PDF（Ozon 下载）")
    print(f"   ├─ 交货货位标签_按ITEMS顺序.pdf（重排后，贴标用）")
    print(f"   ├─ {inquiry_name}")
    print(f"   └─ {order_name}")
    if box_label_path:
        print(f"   推荐打印: {box_label_path}")
    print(f"==================================================")
    return real_order_id


def _api_err_snippet(data, max_len=400):
    if data is None:
        return ""
    try:
        text = json.dumps(data, ensure_ascii=False)
    except TypeError:
        text = str(data)
    if len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def run_merged_application(
    group,
    drop_off_warehouse_name=DEFAULT_CROSSDOCK_DROP_OFF_NAME,
    drop_off_nearby_max_km=80,
    drop_off_retry_delay=6,
    crossdock_delivery_type="DROPOFF",
    seller_warehouse_id=0,
):
    """
    多集群合并单：按发货方式各创 1 个 order_id（直发仅莫斯科；中转多集群）。
    返回 (export_bundle, error_reason, fail_meta)。
    fail_meta 在部分集群货位失败时含 failed_row_ids / ok_row_ids。
    """
    target_shop = group["shop"]
    shipping_method = group["shipping_method"]
    batch_no = group.get("batch_no", "")
    internal_order_no = group.get("internal_order_no", "")
    batch_or_order = group.get("batch_or_order") or resolve_batch_or_order_from_rows(
        group.get("rows") or []
    )
    shipper = group.get("shipper", "")
    archive_date = group.get("ship_date")
    clusters = group.get("clusters") or []
    prefix = f"{shipping_method}单 {shipper}"

    if target_shop not in SHOP_DATA:
        err = f"未找到店铺【{target_shop}】的 Ozon API 配置"
        print(f"❌ {err}")
        return None, err, None

    headers, proxies = build_shop_session(target_shop)
    client_id = SHOP_DATA[target_shop]["client_id"]

    cluster_label = cluster_folder_label(clusters)
    print("=" * 60)
    print(f"📋 合并单: 发货人={shipper}  单号={batch_or_order}  批次={batch_no}")
    print(f"🏪 店铺: 【{target_shop}】 (Client-Id: {client_id})")
    print(f"🌍 集群 ({len(clusters)}): {', '.join(clusters)}")
    print(f"🚀 发货方式: 【{shipping_method}】")
    print("=" * 60)

    print("\n🔍 [STEP 1] 解析 ITEMS 并按集群汇总...")
    try:
        shipment_items, cluster_final_items, cluster_id_map, total_boxes = (
            prepare_merged_shipment(group, headers, proxies)
        )
    except ValueError as e:
        err = str(e)
        print(f"❌ {err}")
        return None, err, None

    print(
        f"✅ 预计总箱数: {total_boxes}。"
        f"集群草稿: { {k: v for k, v in cluster_final_items.items()} }"
    )

    cluster_infos = build_clusters_info_payload(
        cluster_final_items,
        cluster_id_map,
        cluster_order=cluster_names_in_item_order(shipment_items),
    )
    draft_id = None
    selected_cluster_warehouses = None
    final_dropoff_wh_name = ""
    draft_error = ""

    if shipping_method == "直发":
        draft_id, selected_cluster_warehouses, _, draft_error = (
            create_direct_draft_and_poll_warehouse(
                cluster_final_items,
                cluster_id_map,
                headers,
                proxies,
            )
        )
    elif shipping_method == "中转":
        print("\n📦 [STEP 2+3] 创建【多集群中转】草稿...")
        draft_id, selected_cluster_warehouses, final_dropoff_wh_name, draft_error = (
            create_multi_cluster_crossdock_draft(
                cluster_infos,
                cluster_id_map,
                headers,
                proxies,
                drop_off_warehouse_name=drop_off_warehouse_name,
                drop_off_nearby_max_km=drop_off_nearby_max_km,
                drop_off_retry_delay=drop_off_retry_delay,
                crossdock_delivery_type=crossdock_delivery_type,
                seller_warehouse_id=seller_warehouse_id,
            )
        )
    else:
        err = f"未知发货方式: {shipping_method}"
        print(f"❌ {err}")
        return None, err, None

    if not draft_id or not selected_cluster_warehouses:
        err = draft_error or f"{prefix} 草稿/算仓失败"
        print(f"❌ {err}")
        return None, err, None

    time.sleep(2)

    cluster_count = len(cluster_final_items)
    supply_type = resolve_merged_supply_type(cluster_count, shipping_method)
    selected_for_supply = normalize_selected_cluster_warehouses(
        selected_cluster_warehouses, shipping_method, cluster_count
    )

    print(f"\n🚀 [STEP 4] 方案转正 (v2/draft/supply/create)...")
    print(
        f"   supply_type={supply_type}  "
        f"selected_cluster_warehouses={len(selected_for_supply)} 条"
    )
    supply_payload = {
        "draft_id": int(draft_id),
        "selected_cluster_warehouses": selected_for_supply,
        "supply_type": supply_type,
    }
    supply_res = ozon_post(
        "https://api-seller.ozon.ru/v2/draft/supply/create",
        headers,
        supply_payload,
        proxies,
    )

    print("\n⏳ [STEP 5] 轮询 order_id...")
    status_payload = {"draft_id": draft_id}
    if supply_res.get("operation_id"):
        status_payload["operation_id"] = supply_res["operation_id"]

    real_order_id = None
    while True:
        status_res = ozon_post(
            "https://api-seller.ozon.ru/v2/draft/supply/create/status",
            headers,
            status_payload,
            proxies,
        )
        candidate_order_id = status_res.get("order_id") or 0
        if candidate_order_id and int(candidate_order_id) > 0:
            real_order_id = int(candidate_order_id)
            break
        order_ids = status_res.get("result", {}).get("order_ids", [])
        if order_ids:
            real_order_id = int(order_ids[0])
            break
        current_status = status_res.get("status", "")
        if str(current_status).upper() in (
            "FAILED",
            "DRAFTSUPPLYCREATESTATUSFAILED",
        ):
            err = f"{prefix} 方案确认失败: {_api_err_snippet(status_res)}"
            print(f"❌ {err}")
            return None, err, None
        time.sleep(3)

    print(f"✅ 合并供货单 order_id: {real_order_id}")
    order_dir = build_order_output_dir(
        shipper,
        target_shop,
        batch_or_order,
        archive_date=archive_date,
    )
    method_dir = os.path.join(order_dir, shipping_method)
    os.makedirs(method_dir, exist_ok=True)
    print(f"📁 输出目录: {order_dir}  （本单: {method_dir}）")

    print("\n🔎 [STEP 6] 解析全部 supply_id...")
    id_to_name = {v: k for k, v in cluster_id_map.items()}
    supplies = poll_order_supplies(
        headers,
        proxies,
        real_order_id,
        cluster_id_to_name=id_to_name,
        expected_count=len(cluster_final_items),
    )
    if not supplies:
        err = f"{prefix} 未能获取 supply 列表（order_id={real_order_id}）"
        print(f"❌ {err}")
        return None, err, None

    supply_by_cluster = resolve_supply_by_cluster(
        supplies,
        cluster_final_items,
        cluster_id_map,
        shipping_method,
    )

    missing_clusters = set(cluster_final_items.keys()) - set(
        supply_by_cluster.keys()
    )
    if missing_clusters:
        err = f"{prefix} order 中缺少集群 supply: {sorted(missing_clusters)}"
        print(f"❌ {err}")
        return None, err, None

    for name, sup in sorted(supply_by_cluster.items()):
        print(
            f"   集群 {name}: supply_id={sup['supply_id']} "
            f"mc_id={sup.get('macrolocal_cluster_id')}"
        )

    cargo_ids_by_cluster = {}
    items_by_cluster = defaultdict(list)
    for row in shipment_items:
        items_by_cluster[row.get("cluster") or "未知集群"].append(row)

    for cluster_name in sorted(items_by_cluster.keys()):
        sup = supply_by_cluster.get(cluster_name)
        if not sup:
            err = f"{prefix} 集群 {cluster_name} 无 supply_id"
            print(f"❌ {err}")
            return None, err, None
        supply_id = sup["supply_id"]
        cluster_items = items_by_cluster[cluster_name]
        cluster_boxes = sum(
            r["quantity"] // r["number"]
            for r in cluster_items
            if r["number"] > 0
        )
        print(
            f"\n📦 [STEP 7] 集群 {cluster_name} "
            f"(supply_id={supply_id}, {cluster_boxes} 箱)..."
        )
        ok, ordered_ids, cargo_err = submit_cargoes_for_supply(
            supply_id,
            cluster_items,
            headers,
            proxies,
            order_id=real_order_id,
            batch_no=batch_no,
            shop=target_shop,
        )
        if not ok:
            ok_items_by_cluster = {
                name: items_by_cluster[name]
                for name in cargo_ids_by_cluster
            }
            fail_meta = build_cargo_failure_meta(
                cluster_name, cluster_items, ok_items_by_cluster
            )
            err = format_cluster_cargo_failure(
                cluster_name, cluster_items, cargo_err=cargo_err
            )
            print(f"❌ {err}")
            if fail_meta.get("ok_row_ids"):
                try:
                    updated = mark_rows_shipment_applied(fail_meta["ok_row_ids"])
                    print(
                        f"✅ 已回写已成功集群发货状态，更新 {updated} 行 "
                        f"（唯一ID={','.join(fail_meta['ok_row_ids'][:10])}"
                        f"{'...' if len(fail_meta['ok_row_ids']) > 10 else ''}）"
                    )
                except Exception as e:
                    print(f"⚠️ 部分成功行回写数据库失败: {e}")
            return None, err, fail_meta
        cargo_ids_by_cluster[cluster_name] = ordered_ids
        time.sleep(1.5)

    try:
        box_label_path, _, cluster_pdf_bytes = export_merged_labels_and_excel(
            method_dir,
            shipment_items,
            cargo_ids_by_cluster,
            supply_by_cluster,
            headers,
            proxies,
            internal_order_no,
            write_excel=False,
            unified_supply=False,
        )
    except Exception as e:
        err = f"{prefix} 箱唛/Excel 导出失败: {e}"
        print(f"❌ {err}")
        return None, err, None

    print(f"\n{'=' * 50}")
    print("🎉 发货单执行完成")
    print(f"⚙️ 模式: {shipping_method}  集群数: {len(cluster_final_items)}")
    print(f"📦 draft_id: {draft_id}")
    print(f"🆔 order_id: {real_order_id}")
    print(f"🚗 supplies: {[s['supply_id'] for s in supplies]}")
    if final_dropoff_wh_name:
        print(f"🚚 交接仓: {final_dropoff_wh_name}")
    print(f"📦 总箱位: {total_boxes}")
    print(f"📁 本单输出: {method_dir}")
    if box_label_path:
        print(f"   本单箱唛: {box_label_path}")
    print(f"{'=' * 50}")
    bundle = {
        "order_id": real_order_id,
        "shipping_method": shipping_method,
        "shop": target_shop,
        "shipper": shipper,
        "ship_date": archive_date,
        "shipment_items": tag_shipment_items_with_method(
            shipment_items, shipping_method
        ),
        "cargo_ids_by_cluster": cargo_ids_by_cluster,
        "cluster_pdf_bytes": cluster_pdf_bytes,
        "internal_order_no": internal_order_no,
        "rows": group.get("rows") or [],
    }
    return bundle, None, None


def run_group_application(
    group,
    drop_off_warehouse_name=DEFAULT_CROSSDOCK_DROP_OFF_NAME,
):
    """按数据库分组执行发货；返回 (export_bundle, error_reason, fail_meta)。"""
    shipping_method = group.get("shipping_method", "")
    shipper = group.get("shipper", "")
    prefix = f"{shipping_method}单 {shipper}"
    try:
        total_boxes = count_boxes_in_items(group["items"])
    except ValueError as e:
        err = f"{prefix} ITEMS 校验失败: {e}"
        print(f"❌ {err}")
        return None, err, None

    export_bundle, apply_error, fail_meta = run_merged_application(
        group,
        drop_off_warehouse_name=drop_off_warehouse_name,
    )
    if not export_bundle:
        err = apply_error or f"{prefix} 申请失败"
        return None, err, fail_meta

    order_id = export_bundle["order_id"]
    try:
        updated = mark_rows_shipment_applied(group.get("rows", []))
        print(
            f"✅ 已回写数据库发货状态（已申请），"
            f"order_id={order_id}，更新 {updated} 行"
        )
    except Exception as e:
        print(
            f"⚠️ order_id={order_id} 已成功，但回写数据库失败: {e}"
        )

    if total_boxes > MAX_BOXES_PER_SUPPLY_ORDER:
        print(
            f"\n✅ 本单共 {total_boxes} 箱，"
            f"order_id={order_id}（各 supply 独立运输货位）"
        )
    return export_bundle, None, None


def run_resume_cargoes_only(
    target_shop,
    supply_id,
    order_id,
    items,
    batch_no,
    internal_order_no="",
    shipper="",
    archive_date=None,
    source_rows=None,
    cluster="",
    batch_or_order=None,
):
    """对已有供货单 activate（若需要）并补提交剩余货位、下载箱唛。"""
    try:
        headers, proxies = build_shop_session(target_shop)
    except KeyError as e:
        print(f"❌ {e}")
        return None

    if not batch_or_order:
        if source_rows:
            batch_or_order = resolve_batch_or_order_from_rows(source_rows)
        else:
            batch_or_order = (str(batch_no).strip() if batch_no else "") or internal_order_no or "未知单号"

    print("=" * 60)
    print("🔄 断点续传：补提交货位 + 下载箱唛")
    print(f"🏪 店铺: {target_shop}  批次: {batch_no}")
    print(f"🆔 order_id: {order_id}  supply_id: {supply_id}")
    print("=" * 60)

    print("\n🔍 正在解析 ITEMS...")
    shipment_items, _, total_boxes = prepare_shipment_items(
        items, headers, proxies
    )
    if not shipment_items:
        print("❌ 无有效 ITEMS，终止。")
        return None
    print(f"✅ 预计总箱数: {total_boxes}")

    cargoes_list = build_cargoes_from_items(shipment_items)
    box_meta = build_box_meta_from_items(shipment_items)
    existing_box_count = count_box_cargoes(supply_id, headers, proxies)
    print(f"📦 当前 supply 已有 {existing_box_count} 个盒子货位")

    transport_cargo_id = ensure_large_supply_transport_setup(
        supply_id,
        total_boxes,
        headers,
        proxies,
        existing_box_count=existing_box_count,
    )

    remaining = cargoes_list[existing_box_count:]
    if remaining:
        if (
            total_boxes > MAX_BOXES_PER_SUPPLY_ORDER
            and not transport_cargo_id
        ):
            print("❌ 运输货位初始化失败。")
            return None
        print(f"📦 待补提交 {len(remaining)} 箱")
        cargoes_ok, _, _ = submit_cargoes_in_batches(
            supply_id,
            remaining,
            headers,
            proxies,
            append_only=True,
            order_id=order_id,
            batch_no=batch_no,
            shop=target_shop,
            already_submitted=existing_box_count,
            full_total=total_boxes,
        )
        if not cargoes_ok:
            print("❌ 补提交货位失败。")
            return None
        time.sleep(1.5)
    else:
        print("✅ 货位已全部存在，直接进入箱唛下载")

    if transport_cargo_id and total_boxes > MAX_BOXES_PER_SUPPLY_ORDER:
        box_cargo_ids = fetch_ordered_box_cargo_ids(
            supply_id, cargoes_list, headers, proxies
        )
        if not bind_boxes_to_transport_cargo(
            supply_id, transport_cargo_id, box_cargo_ids, headers, proxies
        ):
            print("❌ 运输货位绑定失败。")
            return None

    order_dir = build_order_output_dir(
        shipper,
        target_shop,
        batch_or_order,
        archive_date=archive_date,
    )
    os.makedirs(order_dir, exist_ok=True)
    print(f"📁 输出目录: {order_dir}")
    ordered_cargo_ids = fetch_ordered_cargo_ids_by_keys(
        supply_id, cargoes_list, headers, proxies
    )
    box_label_path = download_box_labels(
        supply_id,
        order_dir,
        headers,
        proxies,
        ordered_cargo_ids=ordered_cargo_ids,
        box_meta=box_meta,
    )
    if not box_label_path:
        print("❌ 箱唛下载失败。")
        return None

    full_box_meta = enrich_box_meta_with_cargo_ids(
        box_meta, ordered_cargo_ids
    )
    try:
        export_batch_excel_reports(
            order_dir,
            full_box_meta,
            internal_order_no,
            batch_or_order=batch_or_order,
        )
    except Exception as exc:
        print(f"⚠️ Excel 导出失败: {exc}")

    if source_rows:
        try:
            updated = mark_rows_shipment_applied(source_rows)
            print(
                f"✅ 已回写数据库发货状态（已申请），"
                f"order_id={order_id}，更新 {updated} 行"
            )
        except Exception as e:
            print(
                f"⚠️ order_id={order_id} 续传成功，但回写数据库失败: {e}"
            )

    print(f"\n🎉 断点续传完成，输出目录: {order_dir}")
    return order_id


def run_resume_merged_cargoes(
    target_shop,
    order_id,
    items,
    rows,
    batch_no,
    internal_order_no="",
    shipper="",
    archive_date=None,
    clusters=None,
    batch_or_order=None,
):
    """多 supply 断点续传：按 order_id 拉全部 supply，逐集群补货位并合并箱唛。"""
    try:
        headers, proxies = build_shop_session(target_shop)
    except KeyError as e:
        print(f"❌ {e}")
        return None

    if not batch_or_order:
        batch_or_order = resolve_batch_or_order_from_rows(rows)

    clusters = clusters or sorted(
        {(item.get("cluster") or "").strip() or "未知集群" for item in items}
    )
    print("=" * 60)
    print("🔄 多 supply 断点续传")
    print(f"🏪 店铺: {target_shop}  order_id: {order_id}")
    print(f"🌍 集群: {', '.join(clusters)}")
    print("=" * 60)

    shipment_items, cluster_final_items, cluster_id_map, total_boxes = (
        prepare_merged_shipment(
            {"items": items, "cluster_items": build_cluster_items_map(items)},
            headers,
            proxies,
        )
    )
    if not shipment_items:
        print("❌ 无有效 ITEMS")
        return None
    print(f"✅ 预计总箱数: {total_boxes}")

    id_to_name = {v: k for k, v in cluster_id_map.items()}
    order_detail_res = ozon_post(
        "https://api-seller.ozon.ru/v3/supply-order/get",
        headers,
        {"order_ids": [str(order_id)]},
        proxies,
    )
    supplies = extract_supplies_from_order_detail(
        order_detail_res, cluster_id_to_name=id_to_name
    )
    if not supplies:
        print("❌ 未能获取 supply 列表")
        return None

    is_crossdock = any(
        sup.get("is_crossdock")
        for order in order_detail_res.get("orders") or []
        for sup in order.get("supplies") or []
    )
    shipping_method = "中转" if is_crossdock else "直发"
    supply_by_cluster = resolve_supply_by_cluster(
        supplies,
        cluster_final_items,
        cluster_id_map,
        shipping_method,
    )
    missing = set(cluster_final_items.keys()) - set(supply_by_cluster.keys())
    if missing:
        print(f"❌ order 中缺少集群 supply: {missing}")
        return None

    cargo_ids_by_cluster = {}
    items_by_cluster = defaultdict(list)
    for row in shipment_items:
        items_by_cluster[row.get("cluster") or "未知集群"].append(row)

    for cluster_name in sorted(items_by_cluster.keys()):
        sup = supply_by_cluster.get(cluster_name)
        if not sup:
            print(f"❌ 集群 {cluster_name} 无 supply_id")
            return None
        supply_id = sup["supply_id"]
        cluster_items = items_by_cluster[cluster_name]
        cargoes_list = build_cargoes_from_items(cluster_items)
        existing_box_count = count_box_cargoes(supply_id, headers, proxies)
        cluster_boxes = len(cargoes_list)
        print(
            f"\n📦 集群 {cluster_name} supply_id={supply_id} "
            f"已有 {existing_box_count}/{cluster_boxes} 箱"
        )

        transport_cargo_id = ensure_large_supply_transport_setup(
            supply_id,
            cluster_boxes,
            headers,
            proxies,
            existing_box_count=existing_box_count,
        )
        remaining = cargoes_list[existing_box_count:]
        if remaining:
            if cluster_boxes > MAX_BOXES_PER_SUPPLY_ORDER and not transport_cargo_id:
                print(f"❌ 集群 {cluster_name} 运输货位初始化失败")
                return None
            ok, _, _ = submit_cargoes_in_batches(
                supply_id,
                remaining,
                headers,
                proxies,
                append_only=True,
                order_id=order_id,
                batch_no=batch_no,
                shop=target_shop,
                already_submitted=existing_box_count,
                full_total=cluster_boxes,
            )
            if not ok:
                print(f"❌ 集群 {cluster_name} 补提交失败")
                return None
            time.sleep(1.5)
        elif cluster_boxes > 0:
            print(f"✅ 集群 {cluster_name} 货位已齐全")

        if transport_cargo_id and cluster_boxes > MAX_BOXES_PER_SUPPLY_ORDER:
            box_cargo_ids = fetch_ordered_box_cargo_ids(
                supply_id, cargoes_list, headers, proxies
            )
            if not bind_boxes_to_transport_cargo(
                supply_id, transport_cargo_id, box_cargo_ids, headers, proxies
            ):
                print(f"❌ 集群 {cluster_name} 运输货位绑定失败")
                return None

        ordered_ids = fetch_ordered_box_cargo_ids(
            supply_id, cargoes_list, headers, proxies
        )
        if len(ordered_ids) != cluster_boxes:
            print(
                f"⚠️ 集群 {cluster_name} cargo_id 数量 "
                f"{len(ordered_ids)}/{cluster_boxes}"
            )
        cargo_ids_by_cluster[cluster_name] = ordered_ids

    order_dir = build_order_output_dir(
        shipper,
        target_shop,
        batch_or_order,
        archive_date=archive_date,
    )
    os.makedirs(order_dir, exist_ok=True)
    print(f"\n📁 输出目录: {order_dir}")

    try:
        export_merged_labels_and_excel(
            order_dir,
            shipment_items,
            cargo_ids_by_cluster,
            supply_by_cluster,
            headers,
            proxies,
            internal_order_no,
            write_excel=True,
            unified_supply=False,
            batch_or_order=batch_or_order,
        )
    except Exception as e:
        print(f"❌ 箱唛/Excel 导出失败: {e}")
        return None

    if rows:
        try:
            updated = mark_rows_shipment_applied(rows)
            print(
                f"✅ 已回写数据库，order_id={order_id}，更新 {updated} 行"
            )
        except Exception as e:
            print(f"⚠️ 回写数据库失败: {e}")

    print(f"\n🎉 多 supply 续传完成，order_id={order_id}")
    return order_id


def main():
    drop_off_warehouse_name = DEFAULT_CROSSDOCK_DROP_OFF_NAME

    print("=" * 60)
    print("📊 从数据库读取今日待发货登记...")
    print("=" * 60)

    try:
        groups = fetch_pending_shipment_groups()
    except Exception as e:
        print(f"❌ 读取数据库失败: {e}")
        return

    if not groups:
        print("✅ 今日无待发货记录（发货状态为空）。")
        return

    bundle_map = defaultdict(list)
    for group in groups:
        bundle_map[group["bundle_key"]].append(group)

    print(
        f"📦 共 {len(groups)} 个申请任务"
        f"（{len(bundle_map)} 个发货人+店铺+单号包，直发/中转分单）\n"
    )
    success_bundles = 0
    task_idx = 0
    for bundle_key, method_groups in sorted(
        bundle_map.items(),
        key=lambda x: (x[0][0], x[0][1], x[0][2], x[0][3]),
    ):
        ship_date, shop, shipper, batch_or_order = bundle_key
        print(f"\n{'#' * 60}")
        print(
            f"▶ 发货包  日期={ship_date}  店铺={shop}  发货人={shipper}  "
            f"单号={batch_or_order}  "
            f"含 {len(method_groups)} 单: "
            f"{', '.join(g['shipping_method'] for g in method_groups)}"
        )
        print(f"{'#' * 60}")

        export_bundles = []
        for group in sorted(
            method_groups,
            key=lambda g: 0 if g["shipping_method"] == "直发" else 1,
        ):
            task_idx += 1
            print(f"\n{'-' * 60}")
            print(
                f"▶ 任务 [{task_idx}/{len(groups)}]  "
                f"方式={group['shipping_method']}  "
                f"单号={group.get('batch_or_order', '')}  "
                f"集群={','.join(group.get('clusters') or [])}  "
                f"批次={group.get('batch_no', '')}"
            )
            print(f"{'-' * 60}")

            try:
                export_bundle, apply_error, _fail_meta = run_group_application(
                    group,
                    drop_off_warehouse_name=drop_off_warehouse_name,
                )
            except Exception as e:
                print(f"❌ {group['shipping_method']}单 {shipper} 执行异常: {e}")
                continue

            if not export_bundle:
                print(
                    f"❌ {apply_error or group['shipping_method'] + '单 ' + shipper + ' 申请失败'}，"
                    f"继续同包下一单。"
                )
                continue

            export_bundles.append(export_bundle)
            if len(export_bundles) < len(method_groups):
                print("⏳ 同包下一单间隔 10 秒...")
                time.sleep(10)

        if not export_bundles:
            print(f"❌ 发货包 {shipper}+{shop}+{batch_or_order} 全部失败。")
            continue

        order_dir = build_order_output_dir(
            shipper, shop, batch_or_order, archive_date=ship_date
        )
        all_rows = []
        for group in method_groups:
            all_rows.extend(group.get("rows") or [])
        internal_order_no = combine_internal_order_nos(all_rows)
        try:
            finalize_combined_exports(
                order_dir,
                export_bundles,
                internal_order_no,
                batch_or_order=batch_or_order,
            )
            success_bundles += 1
            print(
                f"\n✅ 发货包完成: {order_dir}\n"
                f"   总箱唛 + 询价表 + 顺序表已按唯一ID合并"
            )
        except Exception as e:
            print(f"❌ 发货包 {shipper}+{shop}+{batch_or_order} 总表合并失败: {e}")

    print(f"\n{'=' * 60}")
    print(
        f"🏁 全部完成：成功 {success_bundles}/{len(bundle_map)} 个发货人+店铺包"
        f"（共 {len(groups)} 个申请任务）"
    )
    print(f"{'=' * 60}")


def main_resume_cargoes(args):
    ship_date = args.ship_date
    if not ship_date:
        ship_date = datetime.now().date()
    elif isinstance(ship_date, str):
        ship_date = datetime.strptime(ship_date, "%Y-%m-%d").date()
    try:
        rows = fetch_batch_rows(
            args.shop, args.batch, ship_date=ship_date
        )
        items = build_items_from_rows(rows)
        internal_order_no = combine_internal_order_nos(rows)
        shipper = combine_shippers(rows)
        batch_or_order = resolve_batch_or_order_from_rows(rows)
        clusters = sorted(
            {
                (row.get("集群") or "").strip() or "未知集群"
                for row in rows
                if (row.get("集群") or "").strip()
            }
        )
        cluster = cluster_folder_label(clusters) if clusters else ""
    except Exception as e:
        print(f"❌ 读取批次 ITEMS 失败: {e}")
        return

    if args.all_supplies:
        if not args.order_id:
            print("❌ --all-supplies 需要 --order-id")
            return
        order_id = run_resume_merged_cargoes(
            target_shop=args.shop,
            order_id=int(args.order_id),
            items=items,
            rows=rows,
            batch_no=args.batch,
            internal_order_no=internal_order_no,
            shipper=shipper,
            archive_date=ship_date,
            clusters=clusters,
            batch_or_order=batch_or_order,
        )
    elif args.supply_ids:
        if not args.order_id:
            print("❌ --supply-ids 需要 --order-id")
            return
        order_id = run_resume_merged_cargoes(
            target_shop=args.shop,
            order_id=int(args.order_id),
            items=items,
            rows=rows,
            batch_no=args.batch,
            internal_order_no=internal_order_no,
            shipper=shipper,
            archive_date=ship_date,
            clusters=clusters,
            batch_or_order=batch_or_order,
        )
    else:
        if not args.supply_id:
            print("❌ 请指定 --supply-id，或使用 --all-supplies / --supply-ids")
            return
        order_id = run_resume_cargoes_only(
            target_shop=args.shop,
            supply_id=int(args.supply_id),
            order_id=int(args.order_id),
            items=items,
            batch_no=args.batch,
            internal_order_no=internal_order_no,
            shipper=shipper,
            archive_date=ship_date,
            source_rows=rows,
            cluster=cluster,
            batch_or_order=batch_or_order,
        )
    if order_id:
        print(f"✅ 续传成功，order_id={order_id}")


def parse_cli_args():
    parser = argparse.ArgumentParser(description="Ozon FBO 发货全流程")
    parser.add_argument(
        "--resume-cargoes",
        action="store_true",
        help="断点续传：补提交货位+箱唛（单 supply 用 --supply-id；多集群用 --all-supplies）",
    )
    parser.add_argument("--shop", help="店铺，如 OZ-005-RU")
    parser.add_argument("--batch", help="批次号")
    parser.add_argument("--order-id", type=int, help="供货单 order_id")
    parser.add_argument("--supply-id", type=int, help="单个 supply_id（单集群续传）")
    parser.add_argument(
        "--supply-ids",
        help="多个 supply_id，逗号分隔（多集群续传，需配合 --order-id）",
    )
    parser.add_argument(
        "--all-supplies",
        action="store_true",
        help="按 order_id 自动处理全部 supply（多集群合并单续传）",
    )
    parser.add_argument(
        "--ship-date",
        help="运营发货日期 YYYY-MM-DD，默认今天",
    )
    return parser.parse_args()


if __name__ == "__main__":
    cli_args = parse_cli_args()
    if cli_args.resume_cargoes:
        missing = [
            name
            for name, val in [
                ("--shop", cli_args.shop),
                ("--batch", cli_args.batch),
            ]
            if not val
        ]
        if missing:
            print(f"❌ 断点续传缺少参数: {', '.join(missing)}")
            sys.exit(1)
        if not cli_args.all_supplies and not cli_args.supply_ids:
            if not cli_args.order_id or not cli_args.supply_id:
                print(
                    "❌ 单 supply 续传需 --order-id 与 --supply-id；"
                    "多集群请用 --all-supplies --order-id"
                )
                sys.exit(1)
        elif cli_args.all_supplies and not cli_args.order_id:
            print("❌ --all-supplies 需要 --order-id")
            sys.exit(1)
        elif cli_args.supply_ids and not cli_args.order_id:
            print("❌ --supply-ids 需要 --order-id")
            sys.exit(1)
        if cli_args.ship_date:
            cli_args.ship_date = datetime.strptime(
                cli_args.ship_date, "%Y-%m-%d"
            ).date()
        main_resume_cargoes(cli_args)
    else:
        main()